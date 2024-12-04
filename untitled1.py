import tkinter as tk
from tkinter import messagebox, Label, Entry, Button, Frame, ttk, PhotoImage, filedialog,LEFT, RIGHT
from tkinter import *
from PIL import Image, ImageTk
import sqlite3
import random
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime,timedelta
import datetime
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from plyer import notification
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from fpdf import FPDF
import os
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, StandardScaler
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
import tkinter as tk
from tkinter import messagebox, ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
import io 
import configparser 
import sys
img = None
img_tk = None
img_data = None
config_file_path = "theme_config.ini"
class FabricaApp:
    def __init__(self):
        self.conn = sqlite3.connect('user.db')
        self.cursor = self.conn.cursor()
        self.login_attempts = 0
        self.login_window = tk.Tk()
        self.login_window.geometry("600x480")
        self.login_window.title('Giriş Yap')
        self.login_window.config(background="azure")
        self.logo_image = Image.open(
            "C:/Users/admin/Downloads/unnamed (1).png")
        self.logo_image = ImageTk.PhotoImage(self.logo_image)
        self.logo_label = Label(self.login_window, image=self.logo_image)
        self.logo_label.pack(side="top", fill="x")
        self.username_label = Label(
            self.login_window, text='Kullanıcı Adı:', font=("Open Sans", 15))
        self.username_label.pack(padx=50, pady=5)
        self.username_entry = Entry(self.login_window, width=30)
        self.username_entry.pack(padx=20, pady=5)
        self.password_label = Label(
            self.login_window, text='Şifre:', font=("Open Sans", 15))
        self.password_label.pack(padx=20, pady=5)
        self.password_entry = Entry(self.login_window, show="*", width=30)
        self.password_entry.pack(padx=20, pady=5)
        self.space_label = Label(self.login_window, text="", height=5)
        self.space_label.pack()
        self.login_button = Button(
            self.login_window, text='Giriş Yap', command=self.authenticate_user)
        self.login_button.pack()
        with sqlite3.connect('user.db') as conn:
           cursor = conn.cursor()
           cursor.execute("SELECT Makine_Adi FROM makine")
           self.makine_adlari = [row[0] for row in cursor.fetchall()]
           cursor.execute("SELECT Makine_Kodu FROM makine")
           self.makine_kodlari = [row[0] for row in cursor.fetchall()]
           cursor.execute("SELECT name FROM workers")
           self.adlar = [row[0] for row in cursor.fetchall()]
        self.bolumler = ["CNC", "PRES", "KAYNAK"]
        self.data = []
        self.info()
        self.img_path1 = None
        self.img_path2 = None
    def authenticate_user(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        user_data = self.check_user_credentials(username, password)
        if user_data is not None:
            user_authority = user_data[1]
            if user_authority == "A":
                self.login_window.destroy()
                self.create_main_window(user_authority)
            elif user_authority == "B":
                self.create_main_window(user_authority)
        else:
            self.login_attempts += 1
            if self.login_attempts >= 3:
                self.login_window.destroy()
                self.password_reset_page()
            else:
                remaining_attempts = 3 - self.login_attempts
                messagebox.showerror(
                    "Hata", f"Geçersiz kullanıcı adı veya şifre. Kalan deneme hakkınız: {remaining_attempts}")
    def check_user_credentials(self, username, password):
        self.cursor.execute(
            "SELECT username,yetki FROM users WHERE username = ? AND password = ?", (username, password))
        user_data = self.cursor.fetchone()
        return user_data
    def password_reset_page(self):
        self.reset_window = tk.Tk()
        self.reset_window.geometry("400x300")
        self.reset_window.title('Şifremi Unuttum')
        self.email_label = Label(self.reset_window, text='E-posta Adresi:')
        self.email_label.pack(padx=20, pady=10)
        self.email_entry = Entry(self.reset_window)
        self.email_entry.pack(padx=20, pady=10)
        self.send_code_button = Button(
            self.reset_window, text='Kod Gönder', command=self.send_reset_code)
        self.send_code_button.pack(padx=20, pady=10)
        self.reset_code_label = Label(
            self.reset_window, text='Altı Haneli Kod:')
        self.reset_code_label.pack(padx=20, pady=10)
        self.reset_code_entry = Entry(self.reset_window)
        self.reset_code_entry.pack(padx=20, pady=10)
        self.verify_code_button = Button(
            self.reset_window, text='Kodu Doğrula', command=self.verify_reset_code)
        self.verify_code_button.pack(padx=20, pady=10)
        self.reset_code = None
        self.code_sent_time = None
    def open_task_prediction_window(self):
        self.master = tk.Tk()
        self.master.title("Task Completion Time Prediction")
        self.master.geometry("800x600")
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TButton', font=('Arial', 12), padding=10)
        self.style.configure('TLabel', font=('Arial', 12), background='#f0f0f0')
        self.frame = ttk.Frame(self.master, padding="10")
        self.frame.pack(fill=tk.BOTH, expand=True)
        self.setup_model()
        self.create_widgets1()
        self.master.mainloop()
    def setup_model(self):
        conn = sqlite3.connect('user.db')
        query = """
       SELECT task_description, iş_bitis
        FROM tasks"""
        data = pd.read_sql_query(query, conn)
        conn.close()

   # Encode categorical task descriptions
        self.le = LabelEncoder()
        data['task_description'] = self.le.fit_transform(data['task_description'])

   # Prepare features and target variable
        X = data.drop('iş_bitis', axis=1)
        y = data['iş_bitis']
   
   # Split the data into training and testing sets
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

   # Standardize the features
        self.scaler = StandardScaler()
        X_train = self.scaler.fit_transform(X_train)
        X_test = self.scaler.transform(X_test)
   # Build the neural network model
        self.model = Sequential()
        self.model.add(Dense(64, activation='relu', input_shape=(X_train.shape[1],)))
        self.model.add(Dense(32, activation='relu'))
        self.model.add(Dense(1))
   # Compile the model
        self.model.compile(loss='mean_squared_error', optimizer='adam')
   
   # Train the model
        self.history = self.model.fit(X_train, y_train, epochs=100, batch_size=32, validation_split=0.1)
   # Evaluate the model on the test set
        loss = self.model.evaluate(X_test, y_test)
        print("Test Loss:", loss)
   # Create a reverse mapping for labels
        self.reverse_label_mapping = {i: label for i, label in enumerate(self.le.classes_)}
    def create_widgets1(self):
        label = ttk.Label(self.frame, text="Task Description:")
        label.grid(row=0, column=0, pady=10, padx=5, sticky=tk.W)
        self.combobox = ttk.Combobox(self.frame, width=37)
        self.combobox['values'] = [self.reverse_label_mapping[i] for i in range(len(self.reverse_label_mapping))]
        self.combobox.grid(row=0, column=1, pady=10, padx=5)
        predict_button = ttk.Button(self.frame, text="Hesapla", command=self.predict_completion_time)
        predict_button.grid(row=1, column=0, columnspan=2, pady=20)
        self.result_label = ttk.Label(self.frame, text="")
        self.result_label.grid(row=2, column=0, columnspan=2, pady=10)
        plot_button = ttk.Button(self.frame, text="Grafiği göster", command=self.plot_training_graph)
        plot_button.grid(row=3, column=0, columnspan=2, pady=20)
    def predict_completion_time(self):
        new_task_description = self.combobox.get()
        if new_task_description:
            try:
                encoded_description = self.le.transform([new_task_description])
                scaled_description = self.scaler.transform([[encoded_description[0]]])
                prediction = self.model.predict(scaled_description)
                prediction=prediction*100
                prediction = prediction[0][0]
                self.result_label.config(text=f"Predicted Completion Time: {prediction:.2f}")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        else:
            messagebox.showwarning("Input Error", "Please select a task description.")
    def plot_training_graph(self):
        if hasattr(self, 'canvas'):
            self.canvas.get_tk_widget().destroy()
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(self.history.history['loss'], label='Training Loss')
        ax.plot(self.history.history['val_loss'], label='Validation Loss')
        ax.set_title('Training and Validation Loss')
        ax.set_xlabel('Epochs')
        ax.set_ylabel('Loss')
        ax.legend()
        self.canvas = FigureCanvasTkAgg(fig, master=self.frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().grid(row=4, column=0, columnspan=2, pady=10)
    def create_widgets(self):
        self.onarım_kart=tk.Tk()
        self.onarım_kart.title("Bakım Onarım Kartı")
        self.onarım_kart.geometry("600x400")
        # Form elemanlarını oluşturan fonksiyon
        labels = ["Makine Adı", "Makine Kodu", "Bölümü", "Tarih (GG/AA/YYYY)", "Bakımı Yapan", "Planlı mı?", "Gerekçe", "İsim/İmza"]
        for label_text in labels:
            label = ttk.Label(self.onarım_kart, text=label_text)
            label.grid(row=labels.index(label_text), column=0, padx=5, pady=5, sticky="w")
        self.makine_adi_combobox = ttk.Combobox(self.onarım_kart, values=self.makine_adlari)
        self.makine_adi_combobox.grid(row=0, column=1, padx=5, pady=5)
        self.makine_kodu_combobox = ttk.Combobox(self.onarım_kart, values=self.makine_kodlari)
        self.makine_kodu_combobox.grid(row=1, column=1, padx=5, pady=5)
        self.bolum_combobox = ttk.Combobox(self.onarım_kart, values=self.bolumler)
        self.bolum_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.bakim_yapan_combobox = ttk.Combobox(self.onarım_kart, values=self.adlar)
        self.bakim_yapan_combobox.grid(row=4, column=1, padx=5, pady=5)
        self.planli_var = tk.StringVar(value="H")
        planli_checkbox = ttk.Checkbutton(self.onarım_kart, text="", variable=self.planli_var, onvalue="E", offvalue="H")
        planli_checkbox.grid(row=5, column=1, padx=5, pady=5)
        self.gerekce_entry = ttk.Entry(self.onarım_kart)
        self.gerekce_entry.grid(row=6, column=1, padx=5, pady=5)
        self.isim_imza_entry = ttk.Entry(self.onarım_kart)
        self.isim_imza_entry.grid(row=7, column=1, padx=5, pady=5)
        self.tree = ttk.Treeview(self.onarım_kart, columns=labels, show="headings")
        for col in labels:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)  # Sütun genişliği ayarlanabilir
        self.tree.grid(row=8, column=0, columnspan=2, padx=5, pady=5)
        # Butonlar
        self.add_button1 = ttk.Button(self.onarım_kart, text="Kayıt Ekle", command=self.add_record)
        self.add_button1.grid(row=9, column=0, padx=5, pady=5)
        save_button = ttk.Button(self.onarım_kart, text="PDF'ye Kaydet", command=self.save_to_pdf)
        save_button.grid(row=9, column=1, padx=5, pady=5)
    def add_record(self): 
        global data, tree
        makine_adi = self.makine_adi_combobox.get()
        makine_kodu = self.makine_kodu_combobox.get()
        bolum = self.bolum_combobox.get()
        tarih = self.datetime.datetime.now().strftime("%d-%m-%Y")
        bakim_yapan = self.bakim_yapan_combobox.get()
        planli = self.planli_var.get()
        gerekce = self.gerekce_entry.get()
        isim_imza = self.isim_imza_entry.get()
        data.append([makine_adi, makine_kodu, bolum, tarih, bakim_yapan, planli, gerekce, isim_imza])
        self.update_table()
    def update_table():
         global data, tree
         for i in tree.get_children():
             tree.delete(i)
         for row in data:
             tree.insert("", tk.END, values=row)
    def save_to_pdf(self):
        global data
        pdf = FPDF()
        pdf.add_page()
        # Font dosyasının tam yolunu belirtin
        font_path = "C:/Users/admin/Desktop/Yeni klasör (2)/DejaVuSans.ttf"  # Örnek yol (kendi yolunuzu girin)
        if os.path.exists(font_path):
            pdf.add_font('DejaVuSans', '', font_path, uni=True)
            pdf.set_font('DejaVuSans', '', 10)
        else:
            messagebox.showerror("Hata", "Font dosyası bulunamadı: " + font_path)
            return 
        # Başlık
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, "BAKIM-ONARIM KARTI", 0, 1, 'C')
        pdf.ln(10)
        # Tablo başlıkları
        pdf.set_font('Arial', 'B', 12)
        columns = ["Makine Adı", "Makine Kodu", "Bölümü", "Tarih (GG/AA/YYYY)", "Bakımı Yapan", "Planlı mı?", "Gerekçe", "İsim/İmza"]
        column_widths = [30, 30, 30, 30, 30, 20, 40, 30]
        for col, width in zip(columns, column_widths):
            pdf.cell(width, 10, col, 1, 0, 'C')
        pdf.ln()
        # Tablo verileri
        pdf.set_font('Arial', '', 10)
        for row in data:
            for item, width in zip(row, column_widths):
                pdf.cell(width, 10, str(item), 1, 0, 'C')
            pdf.ln()
        # PDF dosyasını kaydet
        pdf_file = "bakim_onarim_karti.pdf"
        pdf.output(pdf_file)
        messagebox.showinfo("Başarılı", f"PDF dosyası kaydedildi: {pdf_file}")
    def send_reset_code(self):
        email = self.email_entry.get().strip()
        if not email:
            messagebox.showerror("Hata", "E-posta adresi boş olamaz.")
            return
        self.reset_code = ''.join(str(random.randint(0, 9)) for _ in range(6))
        self.code_sent_time = time.time()
        try:
            sender_email = "batu.0150@outlook.com"  # E-posta gönderenin adresi
            sender_password = "Batuhan.2018"  # E-posta gönderenin şifresi
            receiver_email = email  # Alıcı e-posta adresi
            message = MIMEMultipart()
            message["From"] = sender_email
            message["To"] = receiver_email
            message["Subject"] = "Şifre Sıfırlama Kodu"
            message.attach(
                MIMEText(f"Şifre sıfırlama kodunuz: {self.reset_code}", "plain"))
            server = smtplib.SMTP("smtp-mail.outlook.com", 587)
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, receiver_email, message.as_string())
            server.quit()
            messagebox.showinfo("Başarılı", "Kod e-posta ile gönderildi.")
        except Exception as e:
            messagebox.showerror(
                "Hata", f"Kod gönderilirken bir hata oluştu: {str(e)}")
    def verify_reset_code(self):
        if not self.reset_code:
            messagebox.showerror("Hata", "Önce kodu göndermelisiniz.")
            return
        entered_code = self.reset_code_entry.get().strip()
        current_time = time.time()
        code_valid_time = 60  # Code valid for 60 seconds
        if current_time - self.code_sent_time > code_valid_time:
            messagebox.showerror(
                "Hata", "Kod süresi doldu. Yeni bir kod gönderin.")
            return
        if entered_code == self.reset_code:
            messagebox.showinfo(
                "Başarılı", "Kod doğrulandı. Şifre sıfırlama işlemine devam edebilirsiniz.")
            self.update_password_page()
        else:
            messagebox.showerror(
                "Hata", "Kod doğrulama başarısız. Lütfen kodu kontrol edin.")
    def update_password_page(self):
        update_password_window = tk.Tk()
        update_password_window.title("Şifre Güncelle")
        update_password_window.geometry("400x300")
        ad_veri_label = tk.Label(update_password_window, text="Kullancı Adı")
        ad_veri_label.pack()
        ad_entry = tk.Entry(update_password_window)
        ad_entry.pack()
        sifre_veri_label = tk.Label(update_password_window, text="Sifre")
        sifre_veri_label.pack()
        sifre_entry = tk.Entry(update_password_window)
        sifre_entry.pack()
        self.update_button = tk.Button(update_password_window, text='Güncelle',
                                       command=lambda: self.password_update(ad_entry.get(), sifre_entry.get()))
        self.update_button.pack()
    def password_update(self, ad, sifre):
        if ad and sifre:  # Kullanıcı adı ve şifre alanları boş değilse
            try:
                self.cursor.execute(
                    "UPDATE users SET password=? WHERE username=?", (sifre, ad))
                self.conn.commit()
                messagebox.showinfo("Başarılı", "Şifre güncellendi!")
            except sqlite3.Error as e:
                messagebox.showerror(
                    "Hata", "Şifre güncelleme işlemi başarısız oldu. Hata: " + str(e))
        else:
            messagebox.showwarning(
                "Uyarı", "Lütfen kullanıcı adı ve şifreyi girin.")
    def export_to_pdf(self):
         selected_item = self.tree1.selection()
         if not selected_item:
             messagebox.showwarning("No selection", "Please select a record to export.")
             return
         item = self.tree1.item(selected_item)
         data = item['values']
         file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
         if file_path:
             self.create_pdf(data, file_path)
             messagebox.showinfo("Success", f"PDF saved to {file_path}")
    def create_pdf(self, data, file_path):
        
         c = canvas.Canvas(file_path, pagesize=A4)
         width, height = A4
         time = datetime.datetime.now().strftime("%d-%m-%Y")
         c.setFont("Helvetica-Bold", 12)
         c.drawString(8 * cm, height - 7 * cm, "ARIZA ONARIM IS EMRI")
         c.setFont("Helvetica", 10)
         c.drawString(15 * cm, height -  8 * cm, "Dokuman No   : FRM.27")
         c.drawString(15 * cm, height - 8.5 * cm, "Sayfa No     : ")
         c.drawString(15 * cm, height - 9 * cm, f"Tarihi : {time}")
         c.drawString(15 * cm, height - 9.5 * cm, "Rev no/tarih : 0")
         c.drawString(2 * cm, height - 11 * cm, "Arizayi Bildiren Birim :")
         c.drawString(6 * cm, height - 11 * cm, data[1])
         c.drawString(2 * cm, height - 11.5 * cm, "Arizayi Bildiren Kisi  :")
         c.drawString(6 * cm, height - 11.5 * cm, data[2])
         c.drawString(2 * cm, height - 12 * cm, "Makina Adi             :")
         c.drawString(6 * cm, height - 12 * cm, data[3])
         c.drawString(13.5 * cm, height - 12 * cm, "Mak. No  :")
         c.drawString(15.5 * cm, height - 12 * cm, data[4])
         c.drawString(2 * cm, height - 12.5 * cm, "Tarih    :")
         c.drawString(6 * cm, height - 12.5 * cm, data[6].split()[0] if data[5] else "")
         c.drawString(13.5 * cm, height - 12.5 * cm, "Saat :")
         c.drawString(15.5 * cm, height - 12.5 * cm, data[6].split()[1] if len(data[6].split()) > 1 else "")
         c.drawString(2 * cm, height - 13 * cm, "Bildirilen Ariza        :")
         c.drawString(6 * cm, height - 13 * cm, data[5])
         c.drawString(13.5 * cm, height - 15 * cm, "Üretim Müdürü ve Imza")
         c.showPage()
         c.save() 
    def find_image_page(self,index):  
        self.scale=1.0
        selected_item = self.tree.selection()
        if selected_item: 
            file_name = self.tree.item(selected_item)["values"][index]  # ImagePath sütununu alır
            file_path = os.path.join(file_name)
            if os.path.isfile(file_path): 
                try: 
                    _, file_extension = os.path.splitext(file_path)
                    file_extension = file_extension.lower()
                    if file_extension in ('.jpg', '.jpeg', '.png'): 
                        self.image = Image.open(file_path)
                        self.resized_image = self.image.resize((300, 250), Image.LANCZOS)
                        self.photo = ImageTk.PhotoImage(self.resized_image)
                        self.new_window = tk.Toplevel(self.main_window)
                        self.new_window.title("Image Viewer")
                        self.new_window.geometry("600x450")
                        self.new_window.resizable(False, False)  
                        self.canvas = tk.Canvas(self.new_window, width=self.resized_image.width, height=self.resized_image.height)
                        self.canvas.pack()
                        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
                        self.canvas.image = self.photo  # referansı tutmak için
                        self.zoom_in_button = tk.Button(self.new_window, text="Yakınlaştır", command=self.zoom_in)
                        self.zoom_in_button.place(x=300, y=400)
                        self.zoom_out_button = tk.Button(self.new_window, text="Uzaklaştır", command=self.zoom_out)
                        self.zoom_out_button.place(x=200, y=400)
                        self.new_window.configure(background=self.bg_color)
                    else: 
                        messagebox.showerror("Error", "Unsupported file format: {}".format(file_name))
                except IOError: 
                    messagebox.showerror("Error", "Hata: {}".format(file_name))
    def zoom_in(self):
        self.scale *= 1.1
        self.update_image()
    def zoom_out(self):
        self.scale /= 1.1
        self.update_image()
    def update_image(self):
        width, height = int(self.resized_image.width * self.scale), int(self.resized_image.height * self.scale)
        self.zoomed_image = self.image.resize((width, height), Image.LANCZOS)
        self.photo = ImageTk.PhotoImage(self.zoomed_image)
        self.canvas.config(width=width, height=height)
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo)
        self.canvas.image = self.photo  # referansı tutmak için
    def load_data5(self): 
        conn = sqlite3.connect('user.db')
        cursor = conn.cursor()
        maintenance_data = pd.read_sql_query('SELECT * FROM makine', conn)
        employee_tasks = pd.read_sql_query('SELECT * FROM makine', conn)
        conn.close()
        return maintenance_data, employee_tasks
    def create_pie_chart(self, active_maintenance, total_machines, ax=None): 
        labels = ['Bakım Yapılan', 'Bakım Yapılmayan']
        sizes = [active_maintenance, total_machines - active_maintenance]
        colors = ['#66b3ff','#ff9999']
        explode = (0.1, 0)  
        if ax is None: 
            fig, ax = plt.subplots()
        ax.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', shadow=True, startangle=140)
        ax.axis('equal')
        ax.set_title('Bakım Durum Grafiği')
        if ax is None: 
            return fig
    def create_bar_chart(self, employee_tasks, ax=None):  
        working_count = employee_tasks[employee_tasks['Calışma_durum'] == 'Çalışıyor'].shape[0]
        break_count = employee_tasks[employee_tasks['Calışma_durum'] == 'Calısmıyor'].shape[0]
        waiting_count = employee_tasks[employee_tasks['Calışma_durum'] == 'Bekliyor'].shape[0] if 'Bekliyor' in employee_tasks['Calışma_durum'].values else 0
        labels = ['Çalışıyor', 'Çalışmıyor', 'Bekliyor']
        sizes = [working_count, break_count, waiting_count]
        colors = ['#66b3ff', '#ff9999', '#99ff99']
        if ax is None: 
            fig, ax = plt.subplots()
        ax.bar(labels, sizes, color=colors)
        ax.set_xlabel('Durum')
        ax.set_ylabel('Sayı')
        ax.set_title('Çalışma Durumu Grafiği')
        if ax is None: 
            return fig 

    def create_main_window(self, user_authority):
       
       self.config = configparser.ConfigParser()
       self.config.read('theme_config.ini')  # Read the config file
       # Set the background color
       self.tema_adi=self.config.get("Theme","selectedtheme")
       self.bg_color = self.config.get('Colors',"colors")  # Default to white if not found
       print(self.bg_color+","+self.tema_adi)
       
       self.main_window = tk.Tk()
       self.main_window.title('Ana Ekran')
       
       # Set the background color
       self.main_window.configure(background=self.bg_color)

       # Get screen dimensions and set the window size and position
       screen_width = self.main_window.winfo_screenwidth()
       screen_height = self.main_window.winfo_screenheight()
       window_width = screen_width
       window_height = screen_height
       window_x = 0
       window_y = 0
       print(screen_height, "+", screen_width)
       self.main_window.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")

       # Create and configure the first Treeview
       self.tree = ttk.Treeview(self.main_window, columns=("Id", "Bakım Yapan", "Makine Adı", "Makine Kodu", "Bölümü", "Bakım Yapıldıgı Tarih", "Bir Sonraki Bakım Tarihi", "Çalışma Durum", "Makine Duruş Saati", "Makine Tekrar Çalısma Saati", "Resim", "Resim1"), show="headings", height=13)
       self.tree.column("#0", width=0, stretch=tk.NO)
       self.tree.heading("Id", text="Id")
       self.tree.heading("Bakım Yapan", text="Bakım Yapan")
       self.tree.heading("Makine Adı", text="Makine Adı")
       self.tree.heading("Makine Kodu", text="Makine Kodu")
       self.tree.heading("Bölümü", text="Bölümü")
       self.tree.heading("Bakım Yapıldıgı Tarih", text="Bakım Yapıldıgı Tarih")
       self.tree.heading("Bir Sonraki Bakım Tarihi", text="Bir Sonraki Bakım Tarihi")
       self.tree.heading("Çalışma Durum", text="Çalışma Durum")
       self.tree.heading("Makine Duruş Saati", text="Makine Duruş Saati")
       self.tree.heading("Makine Tekrar Çalısma Saati", text="Makine Tekrar Çalısma Saati")
       self.tree.heading("Resim", text="Resim")
       self.tree.heading("Resim1", text="Resim1")
       
       # Configure columns
       self.tree.column("Id", width=90)
       self.tree.column("Bakım Yapan", width=90, anchor="center")
       self.tree.column("Bakım Yapıldıgı Tarih", width=90, anchor="center")
       self.tree.column("Makine Adı", width=120, anchor="center")
       self.tree.column("Makine Kodu", width=100, anchor="center")
       self.tree.column("Bölümü", width=90, anchor="center")
       self.tree.column("Bir Sonraki Bakım Tarihi", width=90, anchor="center")
       self.tree.column("Çalışma Durum", width=90, anchor="center")
       self.tree.column("Makine Duruş Saati", width=90, anchor="center")
       self.tree.column("Makine Tekrar Çalısma Saati", width=90, anchor="center")
       self.tree.column("Resim", width=100, anchor="center")
       self.tree.column("Resim1", width=100, anchor="center")

       self.list_records()
       
       # Configure styles
       style = ttk.Style()
       style.configure("Treeview", rowheight=20)
       style.configure("Treeview.Heading", font=("Helvetica", 12))
       self.tree.place(x=370,y=100)
       self.tree.bind("<Button-3>", self.create_popup_menu3)
       
       # Create and configure the second Treeview
       self.tree3 = ttk.Treeview(self.main_window, columns=("Id", "Çalışan Adı", "Yapılan İşlem", "Başlama Tarihi", "Makine Adı", "Makine Kodu", "Çalışma Durumu", "Mola Saati", "İş Bitiş Saati", "Tekrar İşe Başlama", "Geçen Süre"), show="headings", height=13)
       self.tree3.column("0", width=0, stretch=tk.NO)
       self.tree3.heading("Id", text="Id")
       self.tree3.heading("Çalışan Adı", text="Çalışan Adı")
       self.tree3.heading("Yapılan İşlem", text="Yapılan İşlem")
       self.tree3.heading("Başlama Tarihi", text="Başlama Tarihi")
       self.tree3.heading("Makine Adı", text="Makine Kodu")
       self.tree3.heading("Makine Kodu", text="Makine Kodu")
       self.tree3.heading("Çalışma Durumu", text="Çalışma Durumu")
       self.tree3.heading("Mola Saati", text="Mola Saati")
       self.tree3.heading("Tekrar İşe Başlama", text="Tekrar İşe Başlama")
       self.tree3.heading("İş Bitiş Saati", text="İş Bitiş Saati")
       self.tree3.heading("Geçen Süre", text="Geçen Süre")
       
       # Configure columns
       self.tree3.column("Id", width=20)
       self.tree3.column("Çalışan Adı", width=100, anchor="center")
       self.tree3.column("Yapılan İşlem", width=120, anchor="center")
       self.tree3.column("Başlama Tarihi", width=120, anchor="center")
       self.tree3.column("Makine Adı", width=100, anchor="center")
       self.tree3.column("Makine Kodu", width=100, anchor="center")
       self.tree3.column("Çalışma Durumu", width=100, anchor="center")
       self.tree3.column("Mola Saati", width=130, anchor="center")
       self.tree3.column("Tekrar İşe Başlama", width=120, anchor="center")
       self.tree3.column("İş Bitiş Saati", width=120, anchor="center")
       self.tree3.column("Geçen Süre", width=120, anchor="center")
       
       style.configure("Treeview", rowheight=20)
       self.tree3.place(x=370,y=450)
       self.tree3.tag_configure('working', background='#90EE90')
       self.tree3.tag_configure('not_working', background='#FF6347')
       self.tree3.tag_configure('on_break', background='#FFD700')

       # Button and other widgets creation
       if self.tema_adi=="winnative": 
           resim = PhotoImage(file="C:/Users/admin/Downloads/kaynak1.png")
           self.kaynak_butonu = tk.Button(self.main_window, width=350, height=65, image=resim, compound="left", borderwidth=1, relief='flat', command=self.yüz_sayfası)
           self.kaynak_butonu.resim = resim
           self.kaynak_butonu.place(x=(screen_width)/1700, y=100)
           resim1 = PhotoImage(file="C:/Users/admin/Downloads/pres-testere1.png")
           self.testere_butonu = tk.Button(self.main_window, width=350, height=65, image=resim1, compound="left", borderwidth=1, relief="flat", command=self.pres_testere)
           self.testere_butonu.resim1 = resim1
           self.testere_butonu.place(x=(screen_width)/1700, y=200)
           resim2 = PhotoImage(file="C:/Users/admin/Downloads/talaslıimalat1.png")
           self.talaslı_butonu = tk.Button(self.main_window, width=350, height=65, image=resim2, compound="left", borderwidth=1, relief="flat", command=self.talaslı_imalat)
           self.talaslı_butonu.resim2 = resim2
           self.talaslı_butonu.place(x=(screen_width)/1700, y=300)
           resim3 = PhotoImage(file="C:/Users/admin/Downloads/montaj1.png")
           self.montaj_butonu = tk.Button(self.main_window, width=350, height=65, image=resim3, compound="left", borderwidth=1, relief="flat",command=self.montaj_add)
           self.montaj_butonu.resim3 = resim3
           self.montaj_butonu.place(x=(screen_width)/1700, y=400)
           resim4 = PhotoImage(file="C:/Users/admin/Downloads/fabrika-genel1.png")
           self.fabrika_butonu = tk.Button(self.main_window, width=350, height=65, image=resim4, compound="left", borderwidth=1, relief="flat",command=self.fabrika_genel)
           self.fabrika_butonu.resim4 = resim4
           self.fabrika_butonu.place(x=(screen_width)/1700, y=500)
           resim5 = PhotoImage(file="C:/Users/admin/Downloads/arızakontrolu1.png")
           self.arıza_kontrol_butonu = tk.Button(self.main_window, width=350, height=65, image=resim5, compound="left", borderwidth=1, relief="flat", command=self.arıza_kayıt_window)
           self.arıza_kontrol_butonu.resim5 = resim5
           self.arıza_kontrol_butonu.place(x=(screen_width)/1700, y=600)
           resim6 = PhotoImage(file="C:/Users/admin/Downloads/stok ekle.png")
           self.stok_kontrol = tk.Button(self.main_window, width=350, height=65, image=resim6, compound="left", borderwidth=1, relief="flat", command=self.stok_yonetimi)
           self.stok_kontrol.resim6 = resim6
           self.stok_kontrol.place(x=(screen_width)/1700, y=700)
       elif self.tema_adi=="classic":
           resim = PhotoImage(file="C:/Users/admin/Downloads/kaynak2.png")
           self.kaynak_butonu = tk.Button(self.main_window, width=350, height=65, image=resim, compound="left", borderwidth=1, relief='flat', command=self.yüz_sayfası)
           self.kaynak_butonu.resim = resim
           self.kaynak_butonu.place(x=(screen_width)/1700, y=100)
           resim1 = PhotoImage(file="C:/Users/admin/Downloads/pres-testere2.png")
           self.testere_butonu = tk.Button(self.main_window, width=350, height=65, image=resim1, compound="left", borderwidth=1, relief="flat", command=self.pres_testere)
           self.testere_butonu.resim1 = resim1
           self.testere_butonu.place(x=(screen_width)/1700, y=200)
           resim2 = PhotoImage(file="C:/Users/admin/Downloads/talaslıimalat2.png")
           self.talaslı_butonu = tk.Button(self.main_window, width=350, height=65, image=resim2, compound="left", borderwidth=1, relief="flat", command=self.talaslı_imalat)
           self.talaslı_butonu.resim2 = resim2
           self.talaslı_butonu.place(x=(screen_width)/1700, y=300)
           resim3 = PhotoImage(file="C:/Users/admin/Downloads/montaj2.png")
           self.montaj_butonu = tk.Button(self.main_window, width=350, height=65, image=resim3, compound="left", borderwidth=1, relief="flat",command=self.montaj_add)
           self.montaj_butonu.resim3 = resim3
           self.montaj_butonu.place(x=(screen_width)/1700, y=400)
           resim4 = PhotoImage(file="C:/Users/admin/Downloads/fabrika-genel2.png")
           self.fabrika_butonu = tk.Button(self.main_window, width=350, height=65, image=resim4, compound="left", borderwidth=1, relief="flat",command=self.fabrika_genel)
           self.fabrika_butonu.resim4 = resim4
           self.fabrika_butonu.place(x=(screen_width)/1700, y=500)
           resim5 = PhotoImage(file="C:/Users/admin/Downloads/arızakontrol2.png")
           self.arıza_kontrol_butonu = tk.Button(self.main_window, width=350, height=65, image=resim5, compound="left", borderwidth=1, relief="flat", command=self.arıza_kayıt_window)
           self.arıza_kontrol_butonu.resim5 = resim5
           self.arıza_kontrol_butonu.place(x=(screen_width)/1700, y=600)
           resim6 = PhotoImage(file="C:/Users/admin/Downloads/stokyonetimi2.png")
           self.stok_kontrol = tk.Button(self.main_window, width=350, height=65, image=resim6, compound="left", borderwidth=1, relief="flat", command=self.stok_yonetimi)
           self.stok_kontrol.resim6 = resim6
           self.stok_kontrol.place(x=(screen_width)/1700, y=700) 
       else: 
           resim = PhotoImage(file="C:/Users/admin/Downloads/kaynak3.png")
           self.kaynak_butonu = tk.Button(self.main_window, width=350, height=65, image=resim, compound="left", borderwidth=1, relief='flat', command=self.yüz_sayfası)
           self.kaynak_butonu.resim = resim
           self.kaynak_butonu.place(x=(screen_width)/1700, y=100)
           resim1 = PhotoImage(file="C:/Users/admin/Downloads/pres-testere.png")
           self.testere_butonu = tk.Button(self.main_window, width=350, height=65, image=resim1, compound="left", borderwidth=1, relief="flat", command=self.pres_testere)
           self.testere_butonu.resim1 = resim1
           self.testere_butonu.place(x=(screen_width)/1700, y=200)
           resim2 = PhotoImage(file="C:/Users/admin/Downloads/talaslıimalat.png")
           self.talaslı_butonu = tk.Button(self.main_window, width=350, height=65, image=resim2, compound="left", borderwidth=1, relief="flat", command=self.talaslı_imalat)
           self.talaslı_butonu.resim2 = resim2
           self.talaslı_butonu.place(x=(screen_width)/1700, y=300)
           resim3 = PhotoImage(file="C:/Users/admin/Downloads/montaj.png")
           self.montaj_butonu = tk.Button(self.main_window, width=350, height=65, image=resim3, compound="left", borderwidth=1, relief="flat",command=self.montaj_add)
           self.montaj_butonu.resim3 = resim3
           self.montaj_butonu.place(x=(screen_width)/1700, y=400)
           resim4 = PhotoImage(file="C:/Users/admin/Downloads/fabrika-genel.png")
           self.fabrika_butonu = tk.Button(self.main_window, width=350, height=65, image=resim4, compound="left", borderwidth=1, relief="flat",command=self.fabrika_genel)
           self.fabrika_butonu.resim4 = resim4
           self.fabrika_butonu.place(x=(screen_width)/1700, y=500)
           resim5 = PhotoImage(file="C:/Users/admin/Downloads/arızakontrolu.png")
           self.arıza_kontrol_butonu = tk.Button(self.main_window, width=350, height=65, image=resim5, compound="left", borderwidth=1, relief="flat", command=self.arıza_kayıt_window)
           self.arıza_kontrol_butonu.resim5 = resim5
           self.arıza_kontrol_butonu.place(x=(screen_width)/1700, y=600)
           resim6 = PhotoImage(file="C:/Users/admin/Downloads/stok yönetimi.png")
           self.stok_kontrol = tk.Button(self.main_window, width=350, height=65, image=resim6, compound="left", borderwidth=1, relief="flat", command=self.stok_yonetimi)
           self.stok_kontrol.resim6 = resim6
           self.stok_kontrol.place(x=(screen_width)/1700, y=700)  
       export_button = tk.Button(self.main_window, text="CSV'ye Aktar", command=self.export_to_csv)
       export_button.place(x=(screen_width)/1.35, y=400)
       export_button1 = tk.Button(self.main_window, text="CSV'ye Aktar", command=self.export_to_csv2)
       export_button1.place(x=(screen_width)/1.35, y=750)
       is_suresi_tahmin = tk.Button(self.main_window, text='İş Süresi Tahmin Hesaplama', command=self.open_task_prediction_window)
       is_suresi_tahmin.place(x=(screen_width)/1.60, y=750)
       is_suresi_tahmin1 = tk.Button(self.main_window, text='Görsel')
       is_suresi_tahmin1.place(x=(screen_width)/1.40, y=750)
       bakım_sil = tk.Button(self.main_window, text="Bakım Ekle", command=self.makine_ekle)
       bakım_sil.place(x=(screen_width)/1.60, y=400)
       bakım_güncelle = tk.Button(self.main_window, text="Bakım Güncelle", command="#")
       bakım_güncelle.place(x=(screen_width)/1.48, y=400)
       # Image buttons for settings and logout
       self.image1 = PhotoImage(file="C:/Users/admin/Downloads/ayarlar.png")
       self.resized_image1 = self.image1.subsample(10, 10)
       self.resim1_button = tk.Button(self.main_window, width=100, height=60, image=self.resized_image1, compound="left", command=self.ayarlar_page)
       self.resim1_button.pack(side='left')
       self.resim1_button.place(x=1650, y=25)
       self.image2 = PhotoImage(file="C:/Users/admin/Downloads/cıkısyap.png")
       self.resized_image2 = self.image2.subsample(8, 8)
       self.resim2_button = tk.Button(self.main_window, width=100, height=60, image=self.resized_image2, compound="left", relief="flat")
       self.resim2_button.config(bg="mediumblue")
       self.resim2_button.place(x=1800, y=25)
       self.tree.bind("<Button-3>", self.create_popup_menu3)
       self.tree3.bind("<Button-3>", self.create_popup_menu5)
       search_label = tk.Label(self.main_window, text="Aranacak Kelime:")
       search_label.place(x=1300, y=50)
       self.search_entry = tk.Entry(self.main_window)
       self.search_entry.place(x=1400, y=50)
       self.search_entry.bind("<KeyRelease>", self.dynamic_search)
       self.list_records4()
       self.kaynak_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
       self.testere_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
       self.talaslı_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
       self.montaj_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
       self.fabrika_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
       self.arıza_kontrol_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
       self.stok_kontrol.configure(background=self.bg_color,activebackground=self.bg_color)
       self.logo_image = Image.open(
           "C:/Users/admin/Downloads/unnamed (1).png")
       self.logo_image = ImageTk.PhotoImage(self.logo_image)
       self.logo_label = Label(self.main_window, image=self.logo_image)
       maintenance_data, employee_tasks = self.load_data5()
       active_maintenance = maintenance_data[maintenance_data['Calışma_durum'] == 'Çalışıyor'].shape[0]
       total_machines = len(maintenance_data)

# Yeni Figure ve subplotlar oluşturuyoruz (2 satır, 1 sütun)
       fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(5, 10))

# Pie chart ve Bar chart oluşturuluyor
       self.create_pie_chart(active_maintenance, total_machines, ax=ax1)
       self.create_bar_chart(employee_tasks, ax=ax2)
# FigureCanvasTkAgg ile canvas'a ekliyoruz
       canvas = FigureCanvasTkAgg(fig, self.main_window)
       canvas_widget = canvas.get_tk_widget()
       canvas_widget.config(width=300, height=500)
       canvas_widget.place(x=1600, y=130)
       canvas.draw()
    def ayarlar_page(self):
        self.view_ayarlar_window = tk.Toplevel(self.main_window)
        self.view_ayarlar_window.title("Ayarlar")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_width = int(screen_width)
        window_height = int(screen_height)
        window_x = (screen_width - window_width) // 2
        window_y = (screen_height - window_height) // 2
        self.view_ayarlar_window.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.tree15 = ttk.Treeview(self.view_ayarlar_window, columns=("Id", "Kullanıcı Adı", "Sifre", "Yetki"), show="headings")
        self.tree15.heading("Id", text="Id")
        self.tree15.heading("Kullanıcı Adı", text="Kullanıcı Adı")
        self.tree15.heading("Sifre", text="Sifre")
        self.tree15.heading("Yetki", text="Yetki")
        self.tree15.column("Id", width=150, anchor="center")
        self.tree15.column("Kullanıcı Adı", width=300, anchor="center")
        self.tree15.column("Sifre", width=300, anchor="center")
        self.tree15.column("Yetki", width=300, anchor="center")
        self.tree15.pack(pady=20)
        self.tree15.place(x=450, y=100, height=600)
        self.tema_combobox = ttk.Combobox(self.view_ayarlar_window, values=ttk.Style().theme_names())
        self.tema_combobox.set(self.load_theme_config())
        self.tema_combobox.place(x=950, y=22)
        self.tema_degistir_dugmesi = ttk.Button(self.view_ayarlar_window, text="Tema Değiştir", command=lambda: self.tema_degistir(self.tema_combobox.get()))
        self.tema_degistir_dugmesi.place(x=1110, y=20)
        image5 = PhotoImage(file="C:/Users/admin/Downloads/kullancıekle.png")
        self.kullanıcı_ekle = tk.Button(self.view_ayarlar_window, width=350, height=65, image=image5, compound="left", borderwidth=1, relief="flat", command=self.admin_ekle)
        self.kullanıcı_ekle.image = image5
        self.kullanıcı_ekle.place(x=(screen_width)/70, y=250)
        image6 = PhotoImage(file="C:/Users/admin/Downloads/kullanıcısil.png")
        self.kullancı_sil = tk.Button(self.view_ayarlar_window, width=350, height=65, image=image6, compound="left", borderwidth=1, relief="flat",command=self.delete_seleceted_user)
        self.kullancı_sil.image = image6
        self.kullancı_sil.place(x=(screen_width)/70, y=350)
        image7 = PhotoImage(file="C:/Users/admin/Downloads/kullanıcıgüncelle.png")
        self.kullanıcıguncelle = tk.Button(self.view_ayarlar_window, width=350, height=65, image=image7, compound="left", borderwidth=1, relief="flat",command=self.kisi_güncelle)
        self.kullanıcıguncelle.image = image7
        self.kullanıcıguncelle.place(x=(screen_width)/70, y=450)
        image8 = PhotoImage(file="C:/Users/admin/Downloads/renksec.png")
        self.renksec = tk.Button(self.view_ayarlar_window, width=350, height=65, image=image8, compound="left", borderwidth=1, relief="flat")
        self.view_ayarlar_window.configure(background=self.bg_color)
        self.renksec.image = image8
        self.renksec.place(x=(screen_width)/70, y=150)
        self.kullanıcı_ekle.configure(background=self.bg_color,activebackground=self.bg_color)
        self.kullancı_sil.configure(background=self.bg_color,activebackground=self.bg_color)
        self.kullanıcıguncelle.configure(background=self.bg_color,activebackground=self.bg_color)
        self.renksec.configure(background=self.bg_color,activebackground=self.bg_color)
        self.list_records7() 
    def admin_ekle(self):
        self.view_admin_ekle=tk.Tk()
        self.view_admin_ekle.title("Admin Ekle")
        self.view_admin_ekle.geometry("600x400")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        self.admın_ekle_label=tk.Label(self.view_admin_ekle,text="Kullanıcı Ad:")
        self.admın_ekle_label.place(x=screen_width / 8.17, y=0)
        self.admin_ekle_entry=tk.Entry(self.view_admin_ekle)
        self.admin_ekle_entry.place(x=screen_width / 8.17, y=25)
        self.sifre_label=tk.Label(self.view_admin_ekle,text="Sifre:")
        self.sifre_label.place(x=screen_width / 20, y=50)
        self.sifre=tk.Entry(self.view_admin_ekle)
        self.sifre.place(x=screen_width / 8.17, y=53)
        self.yetki_label=tk.Label(self.view_admin_ekle)
        self.yetki_label.place(x=screen_width / 19, y=75)
        self.yetki_combobox=ttk.Combobox(self.view_admin_ekle,values=["A","B"])
        self.yetki_combobox.place(x=screen_width / 8.17, y=75)
        resim5 = tk.PhotoImage(file="C:/Users/admin/Downloads/EKLE ARIZA.png")
        self.add_button2 = tk.Button(self.view_admin_ekle, width=350, height=65, image=resim5,
                                    compound="left", borderwidth=1, relief="flat", command=self.add_admin)
        self.add_button2.resim5 = resim5
        self.add_button2.place(x=150, y=130)
        self.view_admin_ekle.configure(background=self.bg_color)
    def add_admin(self):
        admin_ekle=self.admin_ekle_entry.get()
        sifre=self.sifre.get()
        yetki=self.yetki.get()
        try:
           conn=sqlite.connect("user.db")
           cursor=conn.cursor()
           conn = sqlite3.connect("user.db")
           cursor = conn.cursor()
           cursor.execute("""INSERT INTO makine (username,password,yetki) VALUES (?, ?, ?)""",
                           (admin_ekle,sifre,yetki))
        except Exception as e:
            messagebox.showerror("Hata", f"Eklenirken hata oluştu: {e}")
    def tema_degistir(self, tema_adi):
        stil = ttk.Style()
        stil.theme_use(tema_adi)
        #self.update_theme_settings(tema_adi)
        self.save_theme_config(tema_adi)
        selected_theme = self.load_theme_config()
        stil.theme_use(selected_theme)

    def load_theme_config(self):
        config = configparser.ConfigParser()
        if os.path.exists(config_file_path):
            config.read(config_file_path)
            if 'Theme' in config and 'SelectedTheme' in config['Theme']:
                return config['Theme']['SelectedTheme']
        return ttk.Style().theme_use()
    def save_theme_config(self, tema_adi): 
        colors = {
                    "winnative":"#C0C0C0",  
                    "clam": "#8BC34A", 
                    "default": "#A52A2A",
                    "classic":"#2F4F4F",
                    "vista":"#FFD700",
                    "xpnative":"#4682B4"
        }
        self.bg_color = colors.get(tema_adi, "white")
        self.main_window.configure(background=self.bg_color)
        self.view_ayarlar_window.configure(background=self.bg_color)
        self.kaynak_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
        self.testere_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
        self.talaslı_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
        self.montaj_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
        self.fabrika_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
        self.arıza_kontrol_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
        self.stok_kontrol.configure(background=self.bg_color,activebackground=self.bg_color)
        self.resim2_button.configure(background=self.bg_color,activebackground=self.bg_color)
        self.kullanıcı_ekle.configure(background=self.bg_color,activebackground=self.bg_color)
        self.kullancı_sil.configure(background=self.bg_color,activebackground=self.bg_color)
        self.kullanıcıguncelle.configure(background=self.bg_color,activebackground=self.bg_color)
        self.renksec.configure(background=self.bg_color,activebackground=self.bg_color)
        config = configparser.ConfigParser()
        config['Theme'] = {'SelectedTheme': tema_adi}
        config['Colors']={"Colors":self.bg_color}
        with open(config_file_path, 'w') as config_file: 
            config.write(config_file)
####################3------------------------------------------------#############################################################
    def delete_seleceted_user(self):
        selected_item=self.tree15.selection()
        if selected_item:
            id_to_selected=self.tree15.item(selected_item)["values"][0]
            self.cursor.execute("DELETE FROM users WHERE id=?",(id_to_selected,))
            self.conn.commit()
            messagebox.showinfo("Basarılı","Secili Kisi siindi")
            self.list_records7()
    def kisi_güncelle(self):
        self.kısı_guncelle_window=tk.Toplevel(self.main_window)
        self.kısı_guncelle_window.geometry("600x400")
        self.kısı_guncelle_window.title("Kisi Güncelle")
        self.config.read('theme_config.ini')  # Read the config file
        # Set the background color
        self.bg_color = self.config.get('Colors',"colors")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        self.kisi_label=tk.Label(self.kısı_guncelle_window,text="Kisi Adı")
        self.kisi_label.place(x=screen_width / 13.24, y=0)
        self.kisi_adı_entry=tk.Entry(self.kısı_guncelle_window)
        self.kisi_adı_entry.place(x=screen_width / 8.17, y=0)
        self.password_label=tk.Label(self.kısı_guncelle_window,text="Şifre:")
        self.password_label.place(x=screen_width / 13.24, y=25)
        self.password_entry=tk.Entry(self.kısı_guncelle_window)
        self.password_entry.place(x=screen_width / 8.17, y=25)
        self.yetki_label=tk.Label(self.kısı_guncelle_window,text="Yetki:")
        self.yetki_label.place(x=screen_width / 13.24, y=50)
        self.yetki_combobox=ttk.Combobox(self.kısı_guncelle_window,values=["A","B"])
        self.yetki_combobox.place(x=screen_width / 8.17, y=53)
        resim5 = tk.PhotoImage(file="C:/Users/admin/Downloads/EKLE ARIZA.png")
        self.add_button2 = tk.Button(self.kısı_guncelle_window, width=350, height=65, image=resim5,
                                    compound="left", borderwidth=1, relief="flat", command=self.güncelle_kisi)
        self.add_button2.resim5 = resim5
        self.add_button2.place(x=150, y=130)
        self.kısı_guncelle_window.configure(background=self.bg_color)
        self.add_button2.configure(background=self.bg_color,activebackground=self.bg_color)
    def güncelle_kisi(self):
        name=self.kisi_adı_entry.get()
        password=self.password_entry.get()
        yetki=self.yetki_combobox.get()
        selected_item=self.tree15.selection()
        if selected_item:
            id=self.tree15.item(selected_item)["values"][0]
            conn=sqlite3.connect("user.db")
            cursor.execute("UPDATE users set username=?,password=?,yetki=? WHERE id=?",(name,password,yetki,id))
            conn.commit()
            conn.close()
            messagebox.showinfo("Basarılı","Guncelleme Basarılı")
            self.list_records7()  
    def kisi_güncelle(self):
        self.kısı_guncelle_window = tk.Toplevel(self.main_window)
        self.kısı_guncelle_window.geometry("600x400")
        self.kısı_guncelle_window.title("Kisi Güncelle")
        self.config.read('theme_config.ini')  # Read the config file
        # Set the background color
        self.bg_color = self.config.get('Colors', "colors")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        self.kisi_label = tk.Label(self.kısı_guncelle_window, text="Kisi Adı")
        self.kisi_label.place(x=screen_width / 13.24, y=0)
        self.kisi_adı_entry = tk.Entry(self.kısı_guncelle_window)
        self.kisi_adı_entry.place(x=screen_width / 8.17, y=0)
        self.password_label = tk.Label(self.kısı_guncelle_window, text="Şifre:")
        self.password_label.place(x=screen_width / 13.24, y=25)
        self.password_entry = tk.Entry(self.kısı_guncelle_window)
        self.password_entry.place(x=screen_width / 8.17, y=25)
        self.yetki_label = tk.Label(self.kısı_guncelle_window, text="Yetki:")
        self.yetki_label.place(x=screen_width / 13.24, y=50)
        self.yetki_combobox = ttk.Combobox(self.kısı_guncelle_window, values=["A", "B"])
        self.yetki_combobox.place(x=screen_width / 8.17, y=53)
        resim5 = tk.PhotoImage(file="C:/Users/admin/Downloads/EKLE ARIZA.png")
        self.add_button2 = tk.Button(self.kısı_guncelle_window, width=350, height=65, image=resim5,
                                     compound="left", borderwidth=1, relief="flat", command=self.güncelle_kisi)
        self.add_button2.resim5 = resim5
        self.add_button2.place(x=150, y=130)
        self.kısı_guncelle_window.configure(background=self.bg_color)
        self.add_button2.configure(background=self.bg_color, activebackground=self.bg_color)
    def güncelle_kisi(self):
        name = self.kisi_adı_entry.get()
        password = self.password_entry.get()
        yetki = self.yetki_combobox.get()
        selected_item = self.tree15.selection()
        if selected_item:
            id = self.tree15.item(selected_item)["values"][0]
            conn = sqlite3.connect("user.db")
            cursor = conn.cursor()  # Create cursor object
            cursor.execute("UPDATE users SET username=?, password=?, yetki=? WHERE id=?", (name, password, yetki, id))
            conn.commit()
            conn.close()
            messagebox.showinfo("Basarılı", "Guncelleme Basarılı")
            self.list_records7()      
    def dynamic_search(self, event):
        search_query = self.search_entry.get().lower()
        for child in self.tree.get_children():
            values = self.tree.item(child, "values")
            if any(search_query in str(value).lower() for value in values):
                self.tree.item(child, tags='match')
            else:
                self.tree.item(child, tags='no_match')
        self.tree.tag_configure('match', background='#0078D7')
        self.tree.tag_configure('no_match', background='white')        
    def makine_ekle(self):
        self.makine_ekle_window = tk.Toplevel(self.main_window)
        self.makine_ekle_window.title("Makine Ekle")
        self.makine_ekle_window.geometry("600x400")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        self.bakım_yapan = tk.Label(self.makine_ekle_window, text="Bakım Yapan")
        self.bakım_yapan.place(x=screen_width / 13.24, y=0)
        data = self.fetch_data_from_db()
        self.worker_ids = {name: worker_id for worker_id, name in data}
        worker_names = list(self.worker_ids.keys())
        self.yapan_secim = ttk.Combobox(self.makine_ekle_window, values=worker_names, width=30)
        self.yapan_secim.place(x=screen_width / 8.17, y=0)
        self.makine_ad_label = tk.Label(self.makine_ekle_window, text="Makine Adını Giriniz")
        self.makine_ad_label.place(x=screen_width / 16, y=25)
        self.makine_ad_giris = tk.Entry(self.makine_ekle_window)
        self.makine_ad_giris.place(x=screen_width / 8.17, y=25)
        self.makine_kodu_label = tk.Label(self.makine_ekle_window, text="Makine Kodunu Giriniz")
        self.makine_kodu_label.place(x=screen_width / 20, y=50)
        self.makine_kodu_entry = tk.Entry(self.makine_ekle_window)
        self.makine_kodu_entry.place(x=screen_width / 8.17, y=53)
        self.makine_bölüm = tk.Label(self.makine_ekle_window, text="Makine Bölüm Giriniz")
        self.makine_bölüm.place(x=screen_width / 19, y=75)
        self.makine_combo = ttk.Combobox(self.makine_ekle_window, values=["CNC", "PRES", "KAYNAK", "PLASTİK ENJEKSİYON"])
        self.makine_combo.place(x=screen_width / 8.17, y=75)
        self.makine_durum_label = tk.Label(self.makine_ekle_window, text="Makine Çalışma Durumu")
        self.makine_durum_label.place(x=screen_width / 20, y=100)
        self.makine_durum_entry = ttk.Combobox(self.makine_ekle_window, values=["Çalışıyor", "Çalışmıyor"])
        self.makine_durum_entry.place(x=screen_width / 8.17, y=100)
        resim5 = tk.PhotoImage(file="C:/Users/admin/Downloads/EKLE ARIZA.png")
        self.add_button3 = tk.Button(self.makine_ekle_window, width=350, height=65, image=resim5,
                                    compound="left", borderwidth=1, relief="flat", command=self.add_makine)
        self.add_button3.resim5 = resim5
        self.add_button3.place(x=150, y=130)
        self.add_button3.configure(background=self.bg_color,activebackground=self.bg_color)
        self.resim_button1 = tk.Button(self.makine_ekle_window, text="Resim 1 Seç", command=lambda: self.load_image(1))
        self.resim_button1.place(x=screen_width / 8.17, y=160)
        self.resim_button2 = tk.Button(self.makine_ekle_window, text="Resim 2 Seç", command=lambda: self.load_image(2))
        self.resim_button2.place(x=screen_width / 8.17, y=190)
        self.makine_ekle_window.configure(background=self.bg_color)
    def load_image(self, image_number):
        file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif")])
        if file_path:
            if image_number == 1:
                self.img_path1 = file_path
                print("Image 1 path loaded")
            elif image_number == 2:
                self.img_path2 = file_path
                print("Image 2 path loaded")
    def add_makine(self):
        name = self.yapan_secim.get()
        makine_ad_giris = self.makine_ad_giris.get()
        makine_kodu = self.makine_kodu_entry.get()
        makine_combo = self.makine_combo.get()
        makine_durum = self.makine_durum_entry.get()
        time = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        future_date = None
        if makine_combo in ["CNC", "PRES"]:
            future_date = datetime.datetime.now() + datetime.timedelta(days=365)
        elif makine_combo in ["KAYNAK", "PLASTİK ENJEKSİYON"]:
            future_date = datetime.datetime.now() + datetime.timedelta(days=182)
        future_date_str = future_date.strftime("%d-%m-%Y %H:%M:%S") if future_date else "Unknown"
        try:
            conn = sqlite3.connect("user.db")
            cursor = conn.cursor()
            cursor.execute("""INSERT INTO makine (Bakim_Yapan, Makine_Adi, Makine_Kodu, 
                          Bakim_Yapildigi_Tarih, Bir_Sonraki_Bakim_Tarihi, Bolum, 
                          Calışma_durum, Resim, Resim1) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                           (name, makine_ad_giris, makine_kodu, time, future_date_str,
                            makine_combo, makine_durum, self.img_path1, self.img_path2))
            conn.commit()
            conn.close()
            messagebox.showinfo("Başarılı", "Makine Başarıyla Eklendi")
            self.list_records()
        except Exception as e:
            messagebox.showerror("Hata", f"Eklenirken hata oluştu: {e}")
   #####----------------------------------------------------------------------------------------------#####################
    def makine_guncelle(self):
        self.update_window=tk.Toplevel()
        self.update_window.title("Makine Bilgisi Güncelle")
        self.update_window.geometry("600x400")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_width = screen_width
        window_height = screen_height
        window_x = 0
        window_y = 0
        self.makine_ad_label=tk.Label(self.update_window,text="Makine Adını Giriniz")
        self.makine_ad_label.place(x=(screen_width)/16,y=25)
        self.makine_ad_giris=tk.Entry(self.update_window)
        self.makine_ad_giris.place(x=(screen_width)/8.17,y=25)
        self.makine_kodu_label=tk.Label(self.update_window,text="Makine Kodunu Giriniz")
        self.makine_kodu_label.place(x=(screen_width)/20,y=50)
        self.makine_kodu_entry=tk.Entry(self.update_window)
        self.makine_kodu_entry.place(x=(screen_width)/8.17,y=53)
        self.makine_bölüm=tk.Label(self.update_window,text="Makine Bölüm Giriniz")
        self.makine_bölüm.place(x=(screen_width)/19,y=75)
        self.makine_combo=ttk.Combobox(self.update_window,values=["CNC","PRES","KAYNAK","PLASTİK ENJEKSİYON"])
        self.makine_combo.place(x=(screen_width)/8.17,y=75) 
        # Add your button
        resim5 = PhotoImage(file="C:/Users/admin/Downloads/EKLE ARIZA.png")
        self.add_button4 = tk.Button(self.update_window, width=350, height=65, image=resim5,
                                compound="left", borderwidth=1, relief="flat", command=self.update_güncelle)
        self.add_button4.resim5 = resim5  # Keep a reference to the image
        self.add_button4.place(x=150, y=130)
        self.update_window.configure(background=self.bg_color)
    def update_güncelle(self):
        selected_item=self.tree.selection()
        makine_ad=self.makine_ad_giris.get()
        makine_kodu=self.makine_kodu_entry.get()
        makine_bölüm=self.makine_combo.get()
        if selected_item:
            makine_id=self.tree.item(selected_item)["values"][0]
        try:
            conn=sqlite3.connect("user.db")
            cursor=conn.cursor()
            cursor.execute("UPDATE makine set Makine_Adi=?,Makine_Kodu=?,Bolum=? WHERE Id=?",(makine_ad,makine_kodu,makine_bölüm,makine_id))
            conn.commit()
            conn.close()
            messagebox.showinfo("Başarılı","Güncelleme basarılı")
            self.list_records()
        except Exception as e:
            messagebox.showinfo("Hata",f"Güncelleme oluşurken bir hata ile karsılasıldı:{e}")
    def makine_durdur(self):
        selected_item = self.tree.selection()
        time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if selected_item: 
            makine_id = self.tree.item(selected_item)["values"][0]
            makine_kodu=self.tree.item(selected_item)["values"][3]# Assuming the first column is Id
            try: 
                conn = sqlite3.connect('user.db')
                cursor = conn.cursor()
                cursor.execute("UPDATE makine SET Calışma_durum = ?,makine_durus=? WHERE id = ?", ("Calısmıyor",time, makine_id))
                conn.commit()
                conn.close()
                messagebox.showinfo("Başarılı", "Çalışma durumu başarıyla güncellendi!")
                self.list_records()  
                notification.notify(
                    title='Önemli Uyarı',
                    message=f'Makinelerden:{makine_kodu} kodlu makine durmustur.',
                    app_name='sadasdsad',
                    timeout=10 
                )
                maintenance_data, employee_tasks = self.load_data5()
                active_maintenance = maintenance_data[maintenance_data['Calışma_durum'] == 'Çalışıyor'].shape[0]
                total_machines = len(maintenance_data)
           
           # Yeni Figure ve subplotlar oluşturuyoruz (2 satır, 1 sütun)
                fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(5, 10))
           
           # Pie chart ve Bar chart oluşturuluyor
                self.create_pie_chart(active_maintenance, total_machines, ax=ax1)
                self.create_bar_chart(employee_tasks, ax=ax2)
           
           # Mevcut canvas'ı temizle ve yenisini oluştur
                if hasattr(self, 'canvas_widget'): 
                     self.canvas_widget.destroy()
           
                canvas = FigureCanvasTkAgg(fig, self.main_window)
                self.canvas_widget = canvas.get_tk_widget()
                self.canvas_widget.config(width=300, height=500)
                self.canvas_widget.place(x=1600, y=130)
                canvas.draw()
      
            except Exception as e: 
                messagebox.showerror("Hata", f"Çalışma durumu güncellenirken bir hata oluştu: {e}")
            
    def makine_baslat(self):
        selected_item = self.tree.selection()
        time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if selected_item: 
            makine_id = self.tree.item(selected_item)["values"][0]
            makine_kodu=self.tree.item(selected_item)["values"][3]
            try: 
                conn = sqlite3.connect('user.db')
                cursor = conn.cursor()
                cursor.execute("UPDATE makine SET Calışma_durum = ?,makine_baslat=? WHERE id = ?", ("Calısıyor",time, makine_id))
                conn.commit()
                conn.close()
                messagebox.showinfo("Başarılı", "Çalışma durumu başarıyla güncellendi!")
                self.list_records()  # Refresh the treeview data
                notification.notify(
                    title="Önemli Uyarı",
                    message=f"Makinelerden {makine_kodu} kodlu makine tamir edilip çalışmaya başladı.",
                    app_name='sdfds',
                    timeout=10,
                    )
            except Exception as e: 
                messagebox.showerror("Hata", f"Çalışma durumu güncellenirken bir hata oluştu: {e}")
        
            maintenance_data, employee_tasks = self.load_data5()
            active_maintenance = maintenance_data[maintenance_data['Calışma_durum'] == 'Çalışıyor'].shape[0]
            total_machines = len(maintenance_data)
       
       # Yeni Figure ve subplotlar oluşturuyoruz (2 satır, 1 sütun)
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(5, 10))
       
       # Pie chart ve Bar chart oluşturuluyor
            self.create_pie_chart(active_maintenance, total_machines, ax=ax1)
            self.create_bar_chart(employee_tasks, ax=ax2)
       
       # Mevcut canvas'ı temizle ve yenisini oluştur
            if hasattr(self, 'canvas_widget'): 
                 self.canvas_widget.destroy()
       
            canvas = FigureCanvasTkAgg(fig, self.main_window)
            self.canvas_widget = canvas.get_tk_widget()
            self.canvas_widget.config(width=300, height=500)
            self.canvas_widget.place(x=1600, y=130)
            canvas.draw()
############################------------------------------------------------------------------------------################################
    def fetch_data_from_db1(self):
         conn = sqlite3.connect('user.db')
         cursor = conn.cursor()
         cursor.execute("SELECT id, Makine_Adi FROM makine")
         rows = cursor.fetchall()
         conn.close()
         machine_names = [row[1] for row in rows]  
         return machine_names
############################--------------------------------------------------------------------------------################################
    def stok_yonetimi(self):
        self.stok_yonetimi_view = tk.Toplevel()
        self.stok_yonetimi_view.title('Stok Yönetimi')
        self.fetch_data_from_db1()
        # Ekran boyutunu ve pencereyi ekranın ortasına konumlandırın
        screen_width = self.stok_yonetimi_view.winfo_screenwidth()
        screen_height = self.stok_yonetimi_view.winfo_screenheight()
        window_width = screen_width
        window_height = screen_height
        window_x = 0
        window_y = 0
        print(screen_height, "+", screen_width)
        self.stok_yonetimi_view.geometry(
            f"{window_width}x{window_height}+{window_x}+{window_y}")
        # İlk Treeview'i (self.tree) oluşturun ve yapılandırın
        self.tree5 = ttk.Treeview(self.stok_yonetimi_view, columns=("Id", "Makine Adı", "Bolumu", "Parca Adı", "Miktar"), show="headings", height=15)
        self.tree5.column("#0", width=0, stretch=tk.NO)
        self.tree5.heading("Id", text="Id")
        self.tree5.heading("Makine Adı", text="Makine Adı")
        self.tree5.heading("Bolumu", text="Bolumu")
        self.tree5.heading("Parca Adı", text="Parca Adı")
        self.tree5.heading("Miktar", text="Miktar")
        self.tree5.column("Id", width=250,anchor="center")
        self.tree5.column("Makine Adı", width=250,anchor="center")
        self.tree5.column("Bolumu", width=250,anchor="center")
        self.tree5.column("Parca Adı", width=250,anchor="center")
        self.tree5.column("Miktar", width=250,anchor="center")
        style = ttk.Style()
        style.configure("Treeview")
        style.configure("Treeview.Heading", font=("Helvetica", 12))
        self.tree5.place(x=375,y=100)
        search_label=tk.Label(self.stok_yonetimi_view,text="Aranacak Kelime: ")
        search_label.place(x=1300,y=50)
        self.search_entry=tk.Entry(self.stok_yonetimi_view)
        self.search_entry.place(x=1400,y=50)
        self.search_entry.bind("<KeyRelease>", self.dynamic_search1)
        ######-------------------------------------------------------------------------------------------###########################################################
        self.tree6=ttk.Treeview(self.stok_yonetimi_view,columns=("Id","Stok_Id","Stok_Adı","Kullanılan_Miktar","Tarih"),show="headings",height=15)
        self.tree6.column("#0",width=0,stretch=tk.NO)
        self.tree6.heading("Id",text="Id")
        self.tree6.heading("Stok_Id",text="Stok_Id")
        self.tree6.heading("Stok_Adı",text="Stok_Adı")
        self.tree6.heading("Kullanılan_Miktar",text="Kullanılan_Miktar")
        self.tree6.heading("Tarih",text="Tarih")
        self.tree6.column("Id",width=300,anchor="center")
        self.tree6.column("Stok_Id",width=250,anchor="center")
        self.tree6.column("Stok_Adı",width=250,anchor="center")  
        self.tree6.column("Kullanılan_Miktar",width=250,anchor="center")
        self.tree6.column("Tarih",width=250,anchor="center")
        style=ttk.Style()
        style.configure("Treeview")
        style.configure("Treeview Heading",font=("Helvetica",12))
        self.tree6.place(x=375,y=550)
        ################################3--------------------------------------------------------------------#######################################################
        resim8 = PhotoImage(file="C:/Users/admin/Downloads/stok ekle.png")
        arıza_kontrol_butonu = tk.Button(self.stok_yonetimi_view, width=350, height=65, image=resim8,compound="left", borderwidth=1, relief="flat", command=self.malzeme_ekle)
        arıza_kontrol_butonu.resim8 = resim8
        arıza_kontrol_butonu.place(x=(screen_width)/1080, y=200)
        resim9= PhotoImage(file="C:/Users/admin/Downloads/CSV Aktar.png")
        csv_aktar = tk.Button(self.stok_yonetimi_view, width=350, height=65, image=resim9,compound="left", borderwidth=1, relief="flat", command=self.export_to_csv2)
        csv_aktar.resim9 = resim9
        csv_aktar.place(x=(screen_width)/1080, y=300)
        self.list_records6()
        self.tree5.tag_configure('red', background='#FF6347')  # Tomato
        self.tree5.tag_configure('orange', background='#FFA07A')  # Light Salmon
        self.tree5.tag_configure('yellow', background='#FFD700')  # Gold
        self.tree5.tag_configure('lightgreen', background='#90EE90')  # Light Green
        self.tree5.tag_configure('green', background='#32CD32')  # Lime Green
        self.tree5.bind("<Button-3>", self.create_popup_menu7)
        ########----------------------------------------------------############
        self.tree6.tag_configure('red', background='#FF6347')  # Tomato
        self.tree6.tag_configure('orange', background='#FFA07A')  # Light Salmon
        self.tree6.tag_configure('yellow', background='#FFD700')  # Gold
        self.tree6.tag_configure('lightgreen', background='#90EE90')  # Light Green
        self.tree6.tag_configure('green', background='#32CD32')  # Lime Green
        self.stok_yonetimi_view.configure(background=self.bg_color)
        
        arıza_kontrol_butonu.configure(background=self.bg_color,activebackground=self.bg_color)
        csv_aktar.configure(background=self.bg_color)
        #self.list_records()
        self.list_records7()
        
    def malzeme_ekle(self): 
        self.view_malzeme_ekle_view = tk.Toplevel()  # Corrected the typo here
        self.view_malzeme_ekle_view.title("Malzeme Ekle")
        self.view_malzeme_ekle_view.geometry("600x400")
    # Fetch data from the database
        machine_names = self.fetch_data_from_db1()
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_width = screen_width
        window_height = screen_height
        window_x = 0
        window_y = 0
        self.machine_name_label = tk.Label(self.view_malzeme_ekle_view, text="Malzeme Eklenecek Makine Adı")  # Corrected the attribute here
        self.machine_name_label.pack()
        self.machine_name_combobox = ttk.Combobox(self.view_malzeme_ekle_view, values=machine_names)  # Corrected the attribute here
        self.machine_name_combobox.pack()
        self.machine_name_label=tk.Label(self.view_malzeme_ekle_view,text="Malzeme Eklenecek Bolum Adı")
        self.machine_name_label.pack()
        self.parca_bolumu=ttk.Combobox(self.view_malzeme_ekle_view,values=["CNC","PRES","KAYNAK","PLASTİK ENJEKSİYON"])
        self.parca_bolumu.pack()
        self.parca_adı=tk.Label(self.view_malzeme_ekle_view,text="Malzeme Adı")
        self.parca_adı.pack()
        self.parca_adı_entry=tk.Entry(self.view_malzeme_ekle_view)
        self.parca_adı_entry.pack()
        self.stok_miktarı=tk.Label(self.view_malzeme_ekle_view,text="Stok Miktarı")
        self.stok_miktarı.pack()
        self.stok_miktarı_entry=tk.Entry(self.view_malzeme_ekle_view)
        self.stok_miktarı_entry.pack()
        self.add_button5=tk.Button(self.view_malzeme_ekle_view,text="Ekle",command=self.ekle_stok)
        self.add_button5.pack()
        self.view_malzeme_ekle_view.configure(background=self.bg_color)
    def ekle_stok(self):
        ad=self.machine_name_combobox.get()
        bolum=self.parca_bolumu.get()
        parca_adı=self.parca_adı_entry.get()
        stok_miktarı=self.stok_miktarı_entry.get()
        time=datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        try:
            conn=sqlite3.connect("user.db")
            cursor=conn.cursor()
            cursor.execute("INSERT INTO stok (Makine_Adi,Bolumu,Parca_Adı,Miktar,stok_güncelleme_tarihi) VALUES (?,?,?,?,?)",
                           (ad,bolum,parca_adı,stok_miktarı,time))
            conn.commit()
            conn.close()
            messagebox.showinfo("Basarılı","Stok Ürünü Basarıyla Eklendi")
        except Exception as e:
            messagebox("Hata","Bir Hata ile karsılasıldı:{e}")   
    def stok_azalt(self):
        self.view_stok_guncelle=tk.Toplevel()
        self.view_stok_guncelle.geometry("600x400")
        self.view_stok_guncelle.title("Stok Azalt")
        self.stok_miktarı_label1=tk.Label(self.view_stok_guncelle,text="Kullanılacak Miktarı Giriniz:  ")
        self.stok_miktarı_label1.pack()
        self.stok_miktarı_entry=tk.Entry(self.view_stok_guncelle,textvariable="Stok Miktarı")
        self.stok_miktarı_entry.pack()
        self.guncelle_button=tk.Button(self.view_stok_guncelle,text="Stok Azalt",command=self.azalt_stok)
        self.guncelle_button.pack()
        self.yillik_kullanim_button = tk.Button(self.view_stok_guncelle, text="Yıllık Kullanımı Hesapla", command=self.hesapla_yillik_kullanim)
        self.yillik_kullanim_button.pack()
        self.view_stok_guncelle.configure(background=self.bg_color)
    def azalt_stok(self): 
        stok_guncelle = self.stok_miktarı_entry.get()
        selected_item = self.tree5.selection()
        time = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        if selected_item: 
            stok_id = self.tree5.item(selected_item)["values"][0]
            ilk_miktar = self.tree5.item(selected_item)["values"][4]
            try: 
                stok_guncelle = int(stok_guncelle)  # Ensure stok_guncelle is an integer
                toplam = int(ilk_miktar) - stok_guncelle  # Ensure initial amount is an integer
                if toplam < 0: 
                    raise ValueError("Stok Miktarı Eksi Olamaz")
                conn = sqlite3.connect("user.db")
                cursor = conn.cursor()
            # Update stock amount
                cursor.execute("UPDATE stok SET Miktar=? WHERE Id=?", (toplam, stok_id))
            # Log the stock deduction
                cursor.execute("INSERT INTO stok_log (Stok_Id, Kullanilan_Miktar, Tarih) VALUES (?, ?, ?)", (stok_id, stok_guncelle, time))
                conn.commit()
                conn.close()
                messagebox.showinfo("Başarılı", "Stok Durum başarıyla güncellendi")
                self.list_records6()
            except ValueError as ve: 
                messagebox.showinfo("Hata", f"Hata: {ve}")
            except Exception as e: 
                messagebox.showinfo("Hata", f"Beklenmeyen bir hata oluştu. Lütfen Tekrar Deneyiniz: {e}")
############################--------------------------------------------------------------------------------################################
    def hesapla_yillik_kullanim(self):  
        selected_item = self.tree5.selection()
        if selected_item: 
            stok_id = self.tree5.item(selected_item)["values"][0]
            try: 
                conn = sqlite3.connect("user.db")
                cursor = conn.cursor()
            # Calculate the date one year ago
                one_year_ago = datetime.datetime.now() - datetime.timedelta(days=365)
                one_year_ago_str = one_year_ago.strftime("%Y-%m-%d %H:%M:%S")
            # Query to sum up the usage over the past year
                cursor.execute("SELECT SUM(Kullanilan_Miktar) FROM stok_log WHERE Stok_Id=? AND Tarih>=?", (stok_id, one_year_ago_str))
                result = cursor.fetchone()
                toplam_kullanim = result[0] if result[0] is not None else 0
                conn.close()
                messagebox.showinfo("Yıllık Kullanım", f"Seçilen malzemenin son bir yıl içerisindeki toplam kullanımı: {toplam_kullanim}")
            except Exception as e: 
                messagebox.showinfo("Hata", f"Beklenmeyen bir hata oluştu. Lütfen Tekrar Deneyiniz: {e}")
    def stok_ekle(self):
        self.view_stok_ekle=tk.Toplevel()
        self.view_stok_ekle.geometry("600x400")
        self.view_stok_ekle.title("Stok Ekle")
        self.stok_miktarı_label2=tk.Label(self.view_stok_ekle,text="Eklenecek Ürün Miktarını Giriniz:")
        self.stok_miktarı_label2.pack()
        self.stok_miktarı_entry1=tk.Entry(self.view_stok_ekle)
        self.stok_miktarı_entry1.pack()
        self.ekle_button=tk.Button(self.view_stok_ekle,text="Ekle",command=self.ekle_stok)
        self.ekle_button.pack()
        self.view_stok_ekle.configure(background=self.bg_color)
    def ekle_stok(self):
        stok_miktarı = self.stok_miktarı_entry1.get()
        selected_item = self.tree5.selection()
        time = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        if selected_item:
            try:
                stok_id = self.tree5.item(selected_item)["values"][0]
                ilk_miktar = self.tree5.item(selected_item)["values"][4]
                toplam = int(stok_miktarı) + int(ilk_miktar)
                conn = sqlite3.connect("user.db")
                cursor = conn.cursor()
                cursor.execute("UPDATE stok SET Miktar = ? WHERE Id = ?", (toplam, stok_id))
                conn.commit()
                conn.close()
                messagebox.showinfo("Başarılı", "Stok Başarılı Şekilde Güncellendi.")
                self.list_records6()
            except Exception as e:
                messagebox.showerror("Hata", "Hata Oluştu: " + str(e))
        else:
            messagebox.showwarning("Uyarı", "Lütfen bir öğe seçin.")
    def dynamic_search1(self, event):
        search_query = self.search_entry.get().lower()
        for child in self.tree5.get_children():
            values = self.tree5.item(child, "values")
            if any(search_query in str(value).lower() for value in values):
                self.tree5.item(child, tags='match')
            else:
                self.tree5.item(child, tags="no_match")
            self.tree5s.tag_configure("match", background="#0078D7")
            self.tree5.tag_configure("no_match", background="white")
############################--------------------------------------------------------------------------------################################
    def pres_testere(self):
        self.view_pres_testere = tk.Toplevel()
        self.view_pres_testere.title("Pres-Testere Sayfası")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_width = screen_width
        window_height = screen_height
        window_x = 0
        window_y = 0
        print(screen_height, "+", screen_width)
        self.view_pres_testere.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
        # Create and configure the first Treeview
        self.tree8 = ttk.Treeview(self.view_pres_testere, columns=("Id", "Bakım Yapan", "Makine Adı", "Makine Kodu", "Bölümü", "Bakım Yapıldıgı Tarih", "Bir Sonraki Bakım Tarihi", "Çalışma Durum", "Makine Duruş Saati", "Makine Tekrar Çalısma Saati", "Resim", "Resim1"), show="headings", height=13)
        self.tree8.column("#0", width=0, stretch=tk.NO)
        self.tree8.heading("Id", text="Id")
        self.tree8.heading("Bakım Yapan", text="Bakım Yapan")
        self.tree8.heading("Makine Adı", text="Makine Adı")
        self.tree8.heading("Makine Kodu", text="Makine Kodu")
        self.tree8.heading("Bölümü", text="Bölümü")
        self.tree8.heading("Bakım Yapıldıgı Tarih", text="Bakım Yapıldıgı Tarih")
        self.tree8.heading("Bir Sonraki Bakım Tarihi", text="Bir Sonraki Bakım Tarihi")
        self.tree8.heading("Çalışma Durum", text="Çalışma Durum")
        self.tree8.heading("Makine Duruş Saati", text="Makine Duruş Saati")
        self.tree8.heading("Makine Tekrar Çalısma Saati", text="Makine Tekrar Çalısma Saati")
        self.tree8.heading("Resim", text="Resim")
        self.tree8.heading("Resim1", text="Resim1")
        # Configure columns
        self.tree8.column("Id", width=100)
        self.tree8.column("Bakım Yapan", width=100, anchor="center")
        self.tree8.column("Bakım Yapıldıgı Tarih", width=100, anchor="center")
        self.tree8.column("Makine Adı", width=120, anchor="center")
        self.tree8.column("Makine Kodu", width=100, anchor="center")
        self.tree8.column("Bölümü", width=90, anchor="center")
        self.tree8.column("Bir Sonraki Bakım Tarihi", width=90, anchor="center")
        self.tree8.column("Çalışma Durum", width=90, anchor="center")
        self.tree8.column("Makine Duruş Saati", width=90, anchor="center")
        self.tree8.column("Makine Tekrar Çalısma Saati", width=90, anchor="center")
        self.tree8.column("Resim", width=100, anchor="center")
        self.tree8.column("Resim1", width=100, anchor="center")
        self.list_records9()
        self.tree8.pack(pady=80,padx=120)
        self.tree9 = ttk.Treeview(self.view_pres_testere, columns=("Id", "Birim", "Adı", "Makine Adı", "Makine Kodu",
                                  "Bildirme Saati", "Bildirilen Arıza", "Arızalandıgı Yer", "Arıza Önceligi"), show="headings", height=13)
        self.tree9.heading("Id", text="Id")
        self.tree9.heading("Birim", text="Birim")
        self.tree9.heading("Adı", text="Adı")
        self.tree9.heading("Makine Adı", text="Makine Adı")
        self.tree9.heading("Makine Kodu", text="Makine Kodu")
        self.tree9.heading("Bildirme Saati", text="Bildirme Saati")
        self.tree9.heading("Bildirilen Arıza", text="Bildirilen Arıza")
        self.tree9.heading("Arızalandıgı Yer", text="Arzalandıgı yer")
        self.tree9.heading("Arıza Önceligi", text="Arıza Önceligi")
        self.tree9.column("#0", width=0, stretch=tk.NO)
        self.tree9.column("Id", width=105,anchor="center")
        self.tree9.column("Birim", width=100,anchor="center")
        self.tree9.column("Adı", width=100,anchor="center")
        self.tree9.column("Makine Adı", width=120,anchor="center")
        self.tree9.column("Makine Kodu", width=150,anchor="center")
        self.tree9.column("Bildirme Saati", width=150,anchor="center")
        self.tree9.column("Bildirilen Arıza", width=150,anchor="center")
        self.tree9.column("Arızalandıgı Yer",width=145,anchor="center")
        self.tree9.column("Arıza Önceligi", width=150,anchor="center")
        self.tree9.pack(pady=80,padx=120)
        self.list_records10()
        self.view_pres_testere.configure(background=self.bg_color)
    def talaslı_imalat(self):
        self.view_talasli_imalat=tk.Toplevel()
        self.view_talasli_imalat.title("Talaslı İmalat")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_height=screen_height
        window_width=screen_width
        print(screen_height, "+", screen_width)
        window_x=0
        window_y=0
        self.view_talasli_imalat.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.tree10 = ttk.Treeview(self.view_talasli_imalat, columns=("Id", "Bakım Yapan", "Makine Adı", "Makine Kodu", "Bölümü", "Bakım Yapıldıgı Tarih", "Bir Sonraki Bakım Tarihi", "Çalışma Durum", "Makine Duruş Saati", "Makine Tekrar Çalısma Saati", "Resim", "Resim1"), show="headings", height=13)
        self.tree10.column("#0", width=0, stretch=tk.NO)
        self.tree10.heading("Id", text="Id")
        self.tree10.heading("Bakım Yapan", text="Bakım Yapan")
        self.tree10.heading("Makine Adı", text="Makine Adı")
        self.tree10.heading("Makine Kodu", text="Makine Kodu")
        self.tree10.heading("Bölümü", text="Bölümü")
        self.tree10.heading("Bakım Yapıldıgı Tarih", text="Bakım Yapıldıgı Tarih")
        self.tree10.heading("Bir Sonraki Bakım Tarihi", text="Bir Sonraki Bakım Tarihi")
        self.tree10.heading("Çalışma Durum", text="Çalışma Durum")
        self.tree10.heading("Makine Duruş Saati", text="Makine Duruş Saati")
        self.tree10.heading("Makine Tekrar Çalısma Saati", text="Makine Tekrar Çalısma Saati")
        self.tree10.heading("Resim", text="Resim")
        self.tree10.heading("Resim1", text="Resim1")
        # Configure columns
        self.tree10.column("Id", width=100)
        self.tree10.column("Bakım Yapan", width=100, anchor="center")
        self.tree10.column("Bakım Yapıldıgı Tarih", width=100, anchor="center")
        self.tree10.column("Makine Adı", width=120, anchor="center")
        self.tree10.column("Makine Kodu", width=100, anchor="center")
        self.tree10.column("Bölümü", width=90, anchor="center")
        self.tree10.column("Bir Sonraki Bakım Tarihi", width=90, anchor="center")
        self.tree10.column("Çalışma Durum", width=90, anchor="center")
        self.tree10.column("Makine Duruş Saati", width=90, anchor="center")
        self.tree10.column("Makine Tekrar Çalısma Saati", width=90, anchor="center")
        self.tree10.column("Resim", width=100, anchor="center")
        self.tree10.column("Resim1", width=100, anchor="center")
        self.list_records12()
        self.tree10.pack(pady=80,padx=120)
        self.tree10.pack(pady=90)
        
        self.tree16 = ttk.Treeview(self.view_talasli_imalat, columns=("Id", "Birim", "Adı", "Makine Adı", "Makine Kodu",
                                  "Bildirme Saati", "Bildirilen Arıza", "Arızalandıgı Yer", "Arıza Önceligi"), show="headings", height=13)
        self.tree16.heading("Id", text="Id")
        self.tree16.heading("Birim", text="Birim")
        self.tree16.heading("Adı", text="Adı")
        self.tree16.heading("Makine Adı", text="Makine Adı")
        self.tree16.heading("Makine Kodu", text="Makine Kodu")
        self.tree16.heading("Bildirme Saati", text="Bildirme Saati")
        self.tree16.heading("Bildirilen Arıza", text="Bildirilen Arıza")
        self.tree16.heading("Arızalandıgı Yer", text="Arzalandıgı yer")
        self.tree16.heading("Arıza Önceligi", text="Arıza Önceligi")
        self.tree16.column("#0", width=0, stretch=tk.NO)
        self.tree16.column("Id", width=105,anchor="center")
        self.tree16.column("Birim", width=100,anchor="center")
        self.tree16.column("Adı", width=100,anchor="center")
        self.tree16.column("Makine Adı", width=120,anchor="center")
        self.tree16.column("Makine Kodu", width=150,anchor="center")
        self.tree16.column("Bildirme Saati", width=150,anchor="center")
        self.tree16.column("Bildirilen Arıza", width=150,anchor="center")
        self.tree16.column("Arızalandıgı Yer",width=145,anchor="center")
        self.tree16.column("Arıza Önceligi", width=150,anchor="center")
        self.tree16.pack(pady=80,padx=120)
        
        self.view_talasli_imalat.configure(background=self.bg_color)
    def fabrika_genel(self):
        self.view_fabrika_genel=tk.Toplevel()
        self.view_fabrika_genel.title("Montaj Sayfası")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_height=screen_height
        window_width=screen_width
        print(screen_height, "+", screen_width)
        window_x=0
        window_y=0
        self.view_fabrika_genel.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.tree17=ttk.Treeview(self.view_fabrika_genel,columns=("Id", "Ad", "Birim", "Makine Adı", "Makine Kodu", "Bildirme Saati", "Bildirilen Arıza", "Arızalandıgı Yer", "Arıza Önceliği"))
        self.tree17.heading("Id",text="Id")
        self.tree17.heading("Ad",text="Ad")
        self.tree17.heading("Birim",text="Birim")
        self.tree17.heading("Makine Adı",text="Makine Adı")
        self.tree17.heading("Makine Kodu",text="Makine Kodu")
        self.tree17.heading("Bildirme Saati",text="Bildirme Saati")
        self.tree17.heading("Bildirilen Arıza",text="Bildirilen Arıza")
        self.tree17.heading("Arızalandıgı Yer",text="Arızalandıgı Yer")
        self.tree17.heading("Arıza Önceliği",text="Arıza Önceliği")
        self.tree17.column("#0", width=0, stretch=tk.NO)
        self.tree17.column("Id", width=100, anchor="center")
        self.tree17.column("Ad", width=100, anchor="center")
        self.tree17.column("Birim", width=110, anchor="center")
        self.tree17.column("Makine Adı", width=110, anchor="center")
        self.tree17.column("Makine Kodu", width=110, anchor="center")
        self.tree17.column("Bildirme Saati", width=110, anchor="center")
        self.tree17.column("Bildirilen Arıza", width=110, anchor="center")
        self.tree17.column("Arızalandıgı Yer", width=110, anchor="center")
        self.tree17.column("Arıza Önceliği", width=120, anchor="center")
        self.tree17.pack(pady=90)
        self.tree18 = ttk.Treeview(self.view_fabrika_genel, columns=("Makine Adı", "Makine Kodu", "Bölümü", "Tarih", "Bakım Yapan", "Planlı Mı?", "Gerekçe", "İsim"), show="headings", height=13)
        self.tree18.heading("Makine Adı", text="Makine Adı")
        self.tree18.heading("Makine Kodu", text="Makine Kodu")
        self.tree18.heading("Bölümü", text="Bölümü")
        self.tree18.heading("Tarih", text="Tarih")
        self.tree18.heading("Bakım Yapan", text="Bakım Yapan")
        self.tree18.heading("Planlı Mı?", text="Planlı Mı?")
        self.tree18.heading("Gerekçe", text="Gerekçe")
        self.tree18.heading("İsim", text="İsim")
        self.tree18.column("Makine Adı", width=120, anchor="center")
        self.tree18.column("Makine Kodu", width=120, anchor="center")
        self.tree18.column("Bölümü", width=130, anchor="center")
        self.tree18.column("Tarih", width=130, anchor="center")
        self.tree18.column("Bakım Yapan", width=120, anchor="center")
        self.tree18.column("Planlı Mı?", width=120, anchor="center")
        self.tree18.column("Gerekçe", width=120, anchor="center")
        self.tree18.column("İsim", width=120, anchor="center")
        self.tree18.pack(pady=90)
        resim6 = PhotoImage(file="C:/Users/admin/Downloads/bakım kartı.png")
        self.bakım_kart = tk.Button(self.view_fabrika_genel,width=350,height=65,image=resim6,compound="left",borderwidth=1,relief="flat",command=self.create_form# Assuming you have a method for this
        )
        self.bakım_kart.resim6 = resim6
        self.bakım_kart.place(x=(screen_width)/65, y=175)
        self.view_fabrika_genel.configure(background=self.bg_color)
        self.bakım_kart.configure(background=self.bg_color,activebackground=self.bg_color)
        self.list_records14()
        self.list_records15()
    def montaj_add(self):
        self.view_montaj_add=tk.Toplevel()
        self.view_montaj_add.title("Montaj Sayfası")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_height=screen_height
        window_width=screen_width
        print(screen_height, "+", screen_width)
        window_x=0
        window_y=0
        self.view_montaj_add.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.tree12=ttk.Treeview(self.view_montaj_add,columns=("Id", "Ad", "Birim", "Makine Adı", "Makine Kodu", "Bildirme Saati", "Bildirilen Arıza", "Arızalandıgı Yer", "Arıza Önceliği"))
        self.tree12.heading("Id",text="Id")
        self.tree12.heading("Ad",text="Ad")
        self.tree12.heading("Birim",text="Birim")
        self.tree12.heading("Makine Adı",text="Makine Adı")
        self.tree12.heading("Makine Kodu",text="Makine Kodu")
        self.tree12.heading("Bildirme Saati",text="Bildirme Saati")
        self.tree12.heading("Bildirilen Arıza",text="Bildirilen Arıza")
        self.tree12.heading("Arızalandıgı Yer",text="Arızalandıgı Yer")
        self.tree12.heading("Arıza Önceliği",text="Arıza Önceliği")
        self.tree12.column("#0", width=0, stretch=tk.NO)
        self.tree12.column("Id", width=100, anchor="center")
        self.tree12.column("Ad", width=100, anchor="center")
        self.tree12.column("Birim", width=110, anchor="center")
        self.tree12.column("Makine Adı", width=110, anchor="center")
        self.tree12.column("Makine Kodu", width=110, anchor="center")
        self.tree12.column("Bildirme Saati", width=110, anchor="center")
        self.tree12.column("Bildirilen Arıza", width=110, anchor="center")
        self.tree12.column("Arızalandıgı Yer", width=110, anchor="center")
        self.tree12.column("Arıza Önceliği", width=120, anchor="center")
        self.tree12.pack(pady=90)
        self.tree13 = ttk.Treeview(self.view_montaj_add, columns=("Makine Adı", "Makine Kodu", "Bölümü", "Tarih", "Bakım Yapan", "Planlı Mı?", "Gerekçe", "İsim"), show="headings", height=13)
        self.tree13.heading("Makine Adı", text="Makine Adı")
        self.tree13.heading("Makine Kodu", text="Makine Kodu")
        self.tree13.heading("Bölümü", text="Bölümü")
        self.tree13.heading("Tarih", text="Tarih")
        self.tree13.heading("Bakım Yapan", text="Bakım Yapan")
        self.tree13.heading("Planlı Mı?", text="Planlı Mı?")
        self.tree13.heading("Gerekçe", text="Gerekçe")
        self.tree13.heading("İsim", text="İsim")
        self.tree13.column("Makine Adı", width=120, anchor="center")
        self.tree13.column("Makine Kodu", width=120, anchor="center")
        self.tree13.column("Bölümü", width=130, anchor="center")
        self.tree13.column("Tarih", width=130, anchor="center")
        self.tree13.column("Bakım Yapan", width=120, anchor="center")
        self.tree13.column("Planlı Mı?", width=120, anchor="center")
        self.tree13.column("Gerekçe", width=120, anchor="center")
        self.tree13.column("İsim", width=120, anchor="center")
        self.tree13.pack(pady=90)
        resim6 = PhotoImage(file="C:/Users/admin/Downloads/bakım kartı.png")
        self.bakım_kart = tk.Button(self.view_montaj_add,width=350,height=65,image=resim6,compound="left",borderwidth=1,relief="flat",command=self.create_form# Assuming you have a method for this
        )
        self.bakım_kart.resim6 = resim6
        self.bakım_kart.place(x=(screen_width)/65, y=175)
        self.view_montaj_add.configure(background=self.bg_color)
        self.bakım_kart.configure(background=self.bg_color,activebackground=self.bg_color)
    #####################--------------------------------------------------------------------########################################
    def talep_oluştur(self):
        self.talep_oluştur = tk.Toplevel()
        self.talep_oluştur.title('Talep Oluştur')
        self.talep_oluştur.geometry('600x500')
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_width = screen_width
        window_height = screen_height
        window_x = 0
        window_y = 0
        self.birim_label = tk.Label(self.talep_oluştur, text="Birim Seçiniz")
        self.birim_label.place(x=(screen_width)/13.24,y=0)
        self.birim_combobox = ttk.Combobox(self.talep_oluştur, values=[
                                           "MONTAJ", "KAYNAK", "YÜK", "LOJİSTİK"])
        self.birim_combobox.place(x=(screen_width)/8.17,y=0)
        self.kisi_label = tk.Label(self.talep_oluştur, text="Arızayı Bildiren Kişi")
        self.kisi_label.place(x=(screen_width)/16,y=25)
        self.kisi_entry = tk.Entry(self.talep_oluştur)
        self.kisi_entry.place(x=(screen_width)/8.17,y=25)
        self.bildirilen_arıza = tk.Label(self.talep_oluştur, text="Bildirilen Arıza")
        self.bildirilen_arıza.place(x=(screen_width)/13.24,y=50)
        self.bildirilen_arıza_entry = tk.Entry(self.talep_oluştur)
        self.bildirilen_arıza_entry.place(x=(screen_width)/8.17,y=53)
        self.yer_label = tk.Label(self.talep_oluştur, text="Arıza Yaptıgı Yer")
        self.yer_label.place(x=(screen_width)/14.76,y=75)
        self.yer_entry = ttk.Combobox(self.talep_oluştur, values=[
                                      "ELEKTRONİK", "MEKANİK"])
        self.yer_entry.place(x=(screen_width)/8.17,y=75)
        self.arıza_label = tk.Label(self.talep_oluştur, text="Arıza Önceliği")
        self.arıza_label.place(x=(screen_width)/14.76,y=100)
        self.arıza_onceligi = ttk.Combobox(self.talep_oluştur, values=[
                                           "Yüksek", "Normal", "Az oncelikli"])
        self.arıza_onceligi.place(x=(screen_width)/8.17,y=100)
        resim5 = PhotoImage(file="C:/Users/admin/Downloads/EKLE ARIZA.png")
        self.add_button6= tk.Button(self.talep_oluştur, width=350, height=65, image=resim5,
                                     compound="left", borderwidth=1, relief="flat", command=self.add_talep)
        self.add_button6.resim5 = resim5
        self.add_button6.place(x=150,y=130)
    def arıza_kayıt_window(self):
        self.arıza_kayıt = tk.Toplevel()
        self.arıza_kayıt.title('Arıza Bakım Ekran')
        screen_width = self.arıza_kayıt.winfo_screenwidth()
        screen_height = self.arıza_kayıt.winfo_screenheight()
        window_width = screen_width
        window_height = screen_height
        window_x = 0
        window_y = 0
        self.arıza_kayıt.geometry(
            f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.tree1 = ttk.Treeview(self.arıza_kayıt, columns=("Id", "Birim", "Adı", "Makine Adı", "Makine Kodu",
                                  "Bildirme Saati", "Bildirilen Arıza", "Arızalandıgı Yer", "Arıza Önceligi"), show="headings", height=13)
        self.tree1.heading("Id", text="Id")
        self.tree1.heading("Birim", text="Birim")
        self.tree1.heading("Adı", text="Adı")
        self.tree1.heading("Makine Adı", text="Makine Adı")
        self.tree1.heading("Makine Kodu", text="Makine Kodu")
        self.tree1.heading("Bildirme Saati", text="Bildirme Saati")
        self.tree1.heading("Bildirilen Arıza", text="Bildirilen Arıza")
        self.tree1.heading("Arızalandıgı Yer", text="Arzalandıgı yer")
        self.tree1.heading("Arıza Önceligi", text="Arıza Önceligi")
        self.tree1.column("#0", width=0, stretch=tk.NO)
        self.tree1.column("Id", width=100,anchor="center")
        self.tree1.column("Birim", width=100,anchor="center")
        self.tree1.column("Adı", width=100,anchor="center")
        self.tree1.column("Makine Adı", width=100,anchor="center")
        self.tree1.column("Makine Kodu", width=100,anchor="center")
        self.tree1.column("Bildirme Saati", width=150,anchor="center")
        self.tree1.column("Bildirilen Arıza", width=150,anchor="center")
        self.tree1.column("Arıza Önceligi", width=100,anchor="center")
        self.list_records3()
        self.tree1.pack(pady=80)
        self.tree2 = ttk.Treeview(self.arıza_kayıt, columns=("Id", "Yapan Kişi", "Yapılan İş",
                                  "Makine Adı", "Makine Kodu", "Bitiş Saati", "Arıza Nedeni"), show="headings", height=13)
        self.tree2.heading("Id", text="Id")
        self.tree2.heading("Yapan Kişi", text="Yapılan Kişi")
        self.tree2.heading("Yapılan İş", text="Yapılan İş")
        self.tree2.heading("Makine Adı", text="Makine Adı")
        self.tree2.heading("Makine Kodu", text="Makine Kodu")
        self.tree2.heading("Bitiş Saati", text="Bitiş Saati")
        self.tree2.heading("Arıza Nedeni", text="Arıza Nedeni")
        self.tree2.column("#0", width=0, stretch=tk.NO)
        self.tree2.column("Id", width=150)
        self.tree2.column("Yapan Kişi", width=200,anchor="center")
        self.tree2.column("Yapılan İş", width=150,anchor="center")
        self.tree2.column("Makine Adı", width=150,anchor="center")
        self.tree2.column("Makine Kodu", width=150,anchor="center")
        self.tree2.column("Bitiş Saati", width=150,anchor="center")
        self.tree2.column("Arıza Nedeni", width=150,anchor="center")
        self.tree1.bind("<Button-3>", self.create_popup_menu4)
        self.tree2.bind("<Button-3>", self.create_popup_menu6)
        style = ttk.Style()
        style.configure("Treeview", rowheight=20)
        self.list_record2()
        self.tree2.pack(pady=100)
        self.tree3.tag_configure('green', background='green')
        self.tree3.tag_configure('yellow', background='yellow')
        self.tree3.tag_configure('red', background='red')
        self.arıza_kayıt.configure(background=self.bg_color)
################################--------------------------------------------------------------------------------##################################################
    def add_talep(self):
        selected_item = self.tree.selection()
        if selected_item:
            file_name1 = self.tree.item(selected_item)["values"][2]
            file_name2 = self.tree.item(selected_item)["values"][3]
        birim = self.birim_combobox.get()
        yer = self.yer_entry.get()
        kisi = self.kisi_entry.get()
        ariza = self.bildirilen_arıza_entry.get()
        arıza_onceligi = self.arıza_onceligi.get()
        try:
            conn = sqlite3.connect('user.db')
            cursor = conn.cursor()
            tarih = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute("INSERT INTO ariza_kayit (Birim, Adı,Makine_Adı,Makine_Kodu,Bildirme_Saati,Bildirilen_Arıza,Arıza_bölümü,Acileyet_durumu) VALUES (?, ?, ?, ?, ?, ?,?,?)",
                           (birim, kisi, file_name1, file_name2, ariza, tarih, yer, arıza_onceligi))
            conn.commit()
            conn.close()
            messagebox.showinfo("Başarılı", "Talep Başarıyla Eklendi!")
        except Exception as e:
            messagebox.showerror(
                "Hata", f"Talep eklenirken bir hata oluştu: {e}")
    def export_to_csv(self):
        # Ask the user for a file to save the data to
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path:
            try:
                # Create a new Excel workbook and select the active worksheet
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                # Write the headers
                headers = ["Id", "Yapan Kişi", "Makine Adı", "Makine Kodu",
                           "Bakım Tarihi", "Bir Sonraki Bakım Tarihi","Bölümü","Çalışma Durumu","Makine Durus Saati","Makine Başlatma Saati","Resim1","Resim2"]
                sheet.append(headers)
                # Write the data
                for row_id in self.tree.get_children():
                    row = self.tree.item(row_id)['values']
                    sheet.append(row)
                # Save the workbook
                workbook.save(file_path)
                messagebox.showinfo("Başarılı", "Veriler başarıyla XLSX dosyasına aktarıldı!")
            except Exception as e:
                messagebox.showerror("Hata", f"XLSX dosyasına aktarılırken bir hata oluştu: {e}")
    def export_to_csv1(self):
          file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[
                                                   ("Excel files", "*.xlsx"), ("All files", "*.*")])
          if file_path:
              try:
                  # Create a new Excel workbook and select the active worksheet
                  workbook = openpyxl.Workbook()
                  sheet = workbook.active
                  # Write the headers
                  headers = ["Id", "Çalısan Adı", "Yapılan İşlem", "Baslama Saati",
                             "Makine Adı","Makine Kodu","Çalışma Durumu" ,"Mola Saati", "İş Bitiş Saati","Tekrar İşe Başlama Süresi","Toplam Geçen Süre"]
                  sheet.append(headers)
                  # Write the data
                  for row_id in self.tree3.get_children():
                      row = self.tree3.item(row_id)['values']
                      sheet.append(row)
                  # Save the workbook
                  workbook.save(file_path)
                  messagebox.showinfo(
                      "Başarılı", "Veriler başarıyla XLSX dosyasına aktarıldı!")
              except Exception as e:
                  messagebox.showerror("Hata", f"XLSX dosyasına aktarılırken bir hata oluştu: {e}")
    def export_to_csv2(self): 
        file_path = filedialog.asksaveasfile(defaultextension=".xlsx", filetypes=[("Excel files", ".xlsx"), ("All files", "*.*")])
        if file_path: 
            try: 
                workbook = openpyxl.Workbook()  # Corrected line
                sheet = workbook.active
                headers = ["Id", "Makine_Adi", "Bolumu", "Parca_Adı", "Miktar", "Son Güncellenme Tarihi"]
                sheet.append(headers)
                for row_id in self.tree5.get_children(): 
                    row = self.tree5.item(row_id)['values']
                    sheet.append(row)
                workbook.save(file_path.name)  # Corrected to file_path.name to get the actual file path as string
                messagebox.showinfo("Başarılı", "Veriler başarıyla XLSX dosyasına aktarıldı")
            except Exception as e: 
                messagebox.showerror("Hata", f"XLSX dosyasına aktarılırken bir hata oluştu: {e}")
    def isci_page(self):
        self.view_isci=tk.Toplevel()
        self.view_isci.title("İsci Ekle")
        self.view_isci.geometry("600x400")
        self.add_isci_label=tk.Label(self.view_isci,text="İsci Ekle")
        self.add_isci_label.pack()
        self.add_isci_entry=tk.Entry(self.view_isci)
        self.add_isci_entry.pack()
        self.add_buton=tk.Button(self.view_isci,text="Ekle",command=self.isci_ekle)
        self.add_buton.pack()
        self.view_isci.configure(self)
    def isci_ekle(self):
        isci=self.add_isci_entry.get()
        try:
            # Veritabanı bağlantısını oluştur
            conn = sqlite3.connect('user.db')
            cursor = conn.cursor()
            # Tarih bilgisini al
            # Talep bilgilerini veritabanına ekle
            cursor.execute("INSERT INTO workers (name) VALUES (?)", (isci,))
            conn.commit()
            conn.close()
            messagebox.showinfo("Başarılı", "Talep Başarıyla Eklendi!")
        except Exception as e:
            messagebox.showerror(
                "Hata", f"Talep eklenirken bir hata oluştu: {e}")
    def set_worker_on_break(self): 
        selected_item = self.tree3.selection()
        time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if selected_item: 
            worker_id = self.tree3.item(selected_item)["values"][0]  # Assuming the first column is Id
            try: 
                conn = sqlite3.connect('user.db')
                cursor = conn.cursor()
                cursor.execute("UPDATE tasks SET calısma_durumu = ?,mola_saati=? WHERE id = ?", ("Molada",time, worker_id))
                conn.commit()
                conn.close()
                messagebox.showinfo("Başarılı", "Çalışma durumu başarıyla güncellendi!")
                self.list_records4()  # Refresh the treeview data
            except Exception as e: 
                messagebox.showerror("Hata", f"Çalışma durumu güncellenirken bir hata oluştu: {e}")
    def set_worker_on_break1(self): 
        selected_item = self.tree3.selection()
        time1=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if selected_item: 
            worker_id = self.tree3.item(selected_item)["values"][0]
            zaman=self.tree3.item(selected_item)["values"][3]
            try: 
                conn = sqlite3.connect('user.db')
                cursor = conn.cursor()
                zaman = datetime.datetime.strptime(zaman, "%Y-%m-%d %H:%M:%S")
                time1 = datetime.datetime.strptime(time1, "%Y-%m-%d %H:%M:%S")
                zaman_farkı = zaman - time1
                zaman_farkı_saniye = zaman_farkı.total_seconds()
                zaman_farkı_dakika = abs(zaman_farkı_saniye) / 60
  # zaman_farkı1 değişkeni oluşturulan dakika cinsinden zaman farkını tam bir sayıya yuvarlar
                zaman_farkı1 = round(zaman_farkı_dakika)
                cursor.execute("UPDATE tasks SET calısma_durumu = ?, bitis_saati = ?, iş_bitis = ? WHERE id = ?", ("Çalışmıyor", time1, zaman_farkı1, worker_id))
                conn.commit()
                conn.close()
                messagebox.showinfo("Başarılı", "Çalışma durumu başarıyla güncellendi!")
                self.list_records4()  # Refresh the treeview data
            except Exception as e: 
                messagebox.showerror("Hata", f"Çalışma durumu güncellenirken bir hata oluştu: {e}")
    def set_worker_on_continue(self): 
         selected_item = self.tree3.selection()
         time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
         if selected_item: 
             worker_id = self.tree3.item(selected_item)["values"][0]  # Assuming the first column is Id
             try: 
                 conn = sqlite3.connect('user.db')
                 cursor = conn.cursor()
                 cursor.execute("UPDATE tasks SET calısma_durumu = ?,tekrar_ise_baslama=?  WHERE id = ?", ("Çalısıyor",time, worker_id))
                 conn.commit()
                 conn.close()
                 messagebox.showinfo("Başarılı", "Çalışma durumu başarıyla güncellendi!")
                 self.list_records4()  # Refresh the treeview data
             except Exception as e: 
                 messagebox.showerror("Hata", f"Çalışma durumu güncellenirken bir hata oluştu: {e}")
    def görev_ata(self):
        self.view_gorev_ata=tk.Toplevel()
        self.view_gorev_ata.title("Görev Ata")
        self.view_gorev_ata.geometry("600x400")
        data = self.fetch_data_from_db()
        self.worker_ids = {name: worker_id for worker_id, name in data}
        worker_names = list(self.worker_ids.keys())#çalışan isimlerini çalışan ID'lerine eşleyen bir sözlük (dictionary) oluşturur.
        """  data listesi, fetch_data_from_db fonksiyonundan döner ve (id, name) çiftlerini içerir. Örneğin, data şu şekilde olabilir: [(1, 'Batuhan Şen'), (2, 'Jane Smith')].
             Liste anlayışı (list comprehension) kullanarak, her (id, name) çiftini döner ve bir sözlük oluşturur. Bu sözlükte anahtar (key) olarak çalışan isimleri (name), değer (value) olarak çalışan ID'leri (worker_id) kullanılır.
             self.worker_ids, {name: worker_id for worker_id, name in data} ifadesinin sonucunda şu şekilde bir sözlük olur: {'Batuhan Şen': 1, 'Batuhan Şen': 2}.
        """
        self.yapan_etiket=tk.Label(self.view_gorev_ata,text="Atancak Kişi")
        self.yapan_etiket.place(x=150,y=10)
        self.yapan_label = ttk.Combobox(self.view_gorev_ata, values=worker_names,width=30)
        self.yapan_label.place(x=250,y=10)
        self.task_label=tk.Label(self.view_gorev_ata,text="Atanacak Görev Tanımı")
        self.task_label.place(x=120,y=40)
        self.task_description=tk.Entry(self.view_gorev_ata,width=30)
        self.task_description.place(x=250,y=40)
        self.calısma_durumu_label=tk.Label(self.view_gorev_ata,text="Çalışma Durumu")
        self.calısma_durumu_label.place(x=120,y=70)
        self.calısma_durumu=ttk.Combobox(self.view_gorev_ata,values=["Çalısıyor","Çalışmıyor"],width=30)
        self.calısma_durumu.place(x=250,y=70)
        self.add_buton=tk.Button(self.view_gorev_ata,text="Ekle",command=self.add_atama)
        self.add_buton.place(x=250,y=100)
        self.birim_combobox = ttk.Combobox(self.talep_oluştur, values=["MONTAJ", "KAYNAK", "YÜK", "LOJİSTİK"])
        self.view_gorev_ata.configure(background=self.bg_color)
############################--------------------------------------------------------------------------------################################
    def fetch_data_from_db(self):
        # Veritabanına bağlanma (veritabanı detaylarını değiştirin)
        conn = sqlite3.connect('user.db')
        cursor = conn.cursor()
        # Veriyi çekmek için sorgu çalıştırma
        cursor.execute("SELECT id,name FROM workers")  # Veritabanı şemanıza uygun şekilde değiştirin
        rows = cursor.fetchall()
        # Her satırdan ilk sütunu çıkarma (örneğin adın ilk sütunda olduğunu varsayıyoruz)
        data = []
        for row in rows: 
            try: 
                print("Processing row:", row)
                data.append((row[0], row[1]))
            except IndexError as e: 
                print("IndexError:", e, "in row:", row)
        # Bağlantıyı kapatma
        conn.close()
        return data
    def add_atama(self):
        selected_name = self.yapan_label.get()
        selected_id = self.worker_ids[selected_name]
        yapılacak_is=self.task_description.get()
        selected_item = self.tree1.selection()
        calısma_durumu=self.calısma_durumu.get()
        if selected_item:
            file_name2=self.tree1.item(selected_item)["values"][3]
            file_name3=self.tree1.item(selected_item)["values"][4]
        tarih = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            conn=sqlite3.connect('user.db')
            cursor=conn.cursor()
            cursor.execute("INSERT INTO tasks (worker_id,task_description,start_time,makine_adi,makine_kodu,calısma_durumu) VALUES(?,?,?,?,?,?)",
                           (selected_id,yapılacak_is,tarih,file_name2,file_name3,calısma_durumu))
            conn.commit()
            conn.close()
            messagebox.showinfo("Başarılı", "Talep Başarıyla Eklendi!")
            self.list_records4()
        except Exception as e:
            messagebox.showerror(
                "Hata", f"Talep eklenirken bir hata oluştu: {e}")
    def export_to_pdf1(self):
         selected_item = self.tree2.selection()
         if not selected_item:
             messagebox.showwarning("No selection", "Please select a record to export.")
             return
         item = self.tree2.item(selected_item)
         data = item['values']
         file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
         if file_path:
             self.create_pdf(data, file_path)
             messagebox.showinfo("Success", f"PDF saved to {file_path}")
    def create_pdf1(self, data, file_path): 
        c = canvas.Canvas(file_path, pagesize=A4)
        width, height = A4
        time = datetime.datetime.now().strftime("%d-%m-%Y")
    # Repair Details
        c.drawString(2 * cm, height - 11 * cm, "Yapan Kişi")
        c.drawString(6 * cm, height - 6 * cm, data[1])
        c.drawString(10 * cm, height - 11 * cm, "Yapılan İş")
        c.drawString(6 * cm, height - 6.5 * cm, data[2])
        c.drawString(2 * cm, height - 11.5 * cm, "Makine Adı:")
        c.drawString(6 * cm, height - 6.5 * cm, data[3])
        c.drawString(10 * cm, height - 11.5 * cm, "Makine Kodu:")
        c.drawString(6 * cm, height - 6.5 * cm, data[4])
        c.drawString(2 * cm, height - 12 * cm, "Bitis Saati:")
        c.drawString(6 * cm, height - 6.5 * cm, data[5])
        c.drawString(2 * cm, height - 12.5 * cm, "Arızanın Nedeni:")
        c.drawString(6 * cm, height - 6.5 * cm, data[6])
        c.drawString(2 * cm, height - 14 * cm, "Birim Sorumlusu")
        c.drawString(10 * cm, height - 14 * cm, "Satın Alma ve Üretim Müdürü")
        c.drawString(2 * cm, height - 14.5 * cm, "Adı Soyadı:")
        c.drawString(2 * cm, height - 15 * cm, "İmza:")
        c.drawString(10 * cm, height - 15 * cm, "İmza:")
        c.showPage()
        c.save()
############################--------------------------------------------------------------------------------################################
    def onarım_page(self):
        self.view_onarım_page = tk.Toplevel()
        self.view_onarım_page.title("Onarım Sayfası")
        self.view_onarım_page.geometry("600x400")
        self.yapan_kisi_label = tk.Label(
            self.view_onarım_page, text="Yapan Kişi")
        self.yapan_kisi_label.place(x=120,y=30)
        data = self.fetch_data_from_db()
        self.yapan_kisi=ttk.Combobox(self.view_onarım_page,values=data,width=42)
        self.yapan_kisi.place(x=200,y=30)
        self.yapılan_is_label = tk.Label(
            self.view_onarım_page, text="Yapılan iş")
        self.yapılan_is_label.place(x=120,y=60)
        self.yapılan_is_entry = tk.Entry(self.view_onarım_page, width=45)
        self.yapılan_is_entry.place(x=200,y=60)
        self.arıza_nedeni_label = tk.Label(
            self.view_onarım_page, text="Arıza Nedeni")
        self.arıza_nedeni_label.place(x=120,y=90)
        self.arıza_nedeni_entry = tk.Entry(self.view_onarım_page, width=45)
        self.arıza_nedeni_entry.place(x=200,y=90)
        resim5 = PhotoImage(file="C:/Users/admin/Downloads/EKLE ARIZA.png")
        self.add_button7= tk.Button(self.view_onarım_page, width=350, height=65, image=resim5,
                                         compound="left", borderwidth=1, relief="flat", command=self.add_onarım)
        self.add_button7.resim5 = resim5
        self.add_button7.place(x=150,y=130)
        self.view_onarım_page.configure(background=self.bg_color)
        self.add_button7.configure(background=self.bg_color,activebackground=self.bg_color)
    def add_onarım(self):
        yapan_kisi = self.yapan_kisi.get()
        yapılan_is = self.yapılan_is_entry.get()
        arıza_nedeni = self.arıza_nedeni_entry.get()
        selected_item = self.tree1.selection()
        if selected_item:
            file_name1 = self.tree1.item(selected_item)["values"][3]
            file_name2 = self.tree1.item(selected_item)["values"][4]
        try:
            # Veritabanı bağlantısını oluştur
            conn = sqlite3.connect('user.db')
            cursor = conn.cursor()
            # Tarih bilgisini al
            tarih = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # Talep bilgilerini veritabanına ekle
            cursor.execute("INSERT INTO onarım_kayit (yapan_kisi, yapılan_is, makine_adı, makine_kodu,bitis_saati, arıza_nedeni) VALUES (?, ?, ?, ?, ?, ?)",
                           (yapan_kisi, yapılan_is, file_name1, file_name2, tarih, arıza_nedeni))
            conn.commit()
            conn.close()
            messagebox.showinfo("Başarılı", "Talep Başarıyla Eklendi!")
        except Exception as e:
            messagebox.showerror(
                "Hata", f"Talep eklenirken bir hata oluştu: {e}")
    def info(self):
         self.tasks = [
             "1. Bakım çalışmasına başlamadan önce gerekli iş güvenliği alınmalı ve gerekli olan koruyucu malzemeler kullanılmalıdır.",
             "2. Bakım yapıldığını bildiren uyarı levhası asılmalı.",
             "3. İş eldiveni kullanılmalı.",
             "4. Makinenin elektrik enerjisi kesilmelidir.",
             "5. Makinenin üzerindeki kontrol araç ve bakımlar yapılırken emniyet kemeri takılmalı.",
             "6. Makine çalıştırılarak kaçak kontrol yapılırken hidrolik borulara fazla yaklaşılmamalı ve boruların sağlamlığı kontrol edilmeli.",
             "7. Genel vücut sağlığı gözetilmelidir.",
             "8. Makine ve hidrolik ünitenin ve makinenin komple genel temizliği yapılır.",
             "9. Hidrolik yağ seviyesi kontrol edilir. Eksikse tamamlanır. (Telus 66)",
             "10. Hidrolik tank dip boşluğu filtresi temizlenir.",
             "11. Hidrolik sistem kaçak kontrolü yapılmalı kaçak varsa giderilmeli ve temizlenmeli.",
             "12. Kızaklar kontrol edilmeli ve gerekli ayarlar yapılmalı.",
             "13. Sıvı ve pozisyon sensörlerinin fonksiyonları kontrol edilir.",
             "14. Elektrik panosu ve tesisatı kontrol edilmeli, gevşeklik kontrolü ve temizliği yapılmalı. Aksaklıklar giderilmeli.",
             "15. Elektrik sistemin voltaj, amper ve diğer ölçümleri yapılmalı anomal durum var ise giderilmeli.",
             "16. Genel gevşeklik kontrolü yapılmalı.",
             "17. Kesintime bakımı, makine üzerinde asılı olan ölçüm tablosuna göre yapılır, ölçümler değiştirilir.",
             "18. Bakıma başlanmadan önce tüm dikkatin, aletin güvenliği toplanır ve ilgili alet kapılarına aktarılır. Bakım yapılacak yerde elektrik varsa kesilmelidir.",
             "19. Bakıma başlamadan önce atık oluşan (Absorban malz. Kağıt, vs Maddeler ve atık Yağlar Tehlikeli Atık Mahallinde toplanır."
         ]
         self.checkbox12_vars = []
         self.checkbox36_vars = []
         self.duration_entries = []
    def save_as_pdf(self): 
        font_path = os.path.join(os.path.dirname("C:/Users/admin/Desktop/Yeni klasör (2)/DejaVuSans.ttf"), 'DejaVuSans.ttf')
        pdfmetrics.registerFont(TTFont('DejaVu', font_path))
        pdf = SimpleDocTemplate("bakım_raporu.pdf", pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='TurkishTitle', fontName='DejaVu', fontSize=18))
        styles.add(ParagraphStyle(name='TurkishSubtitle', fontName='DejaVu', fontSize=14))
        styles.add(ParagraphStyle(name='TurkishBody', fontName='DejaVu', fontSize=8))
        title = "BAKIM KARTI"
        subtitle = "6 AYLIK BAKIM PERİYODU İŞ PLANI"
        elements.append(Paragraph(title, styles['TurkishTitle']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(subtitle, styles['TurkishSubtitle']))
        elements.append(Spacer(1, 12))
        table_data = [["Bakım İşlemi", "12", "36", "Süre (dk)"]] 
        for i in range(len(self.tasks)): 
            try: 
                task_data = [
                Paragraph(self.tasks[i], styles['TurkishBody']),
                 "X" if self.checkbox12_vars[i].get() else "",
                 "X" if self.checkbox36_vars[i].get() else "",
                 self.duration_entries[i].get(),]
            except tk.TclError as e: 
                print(f"Error accessing widget: {e}")
            task_data = [
            Paragraph(self.tasks[i], styles['TurkishBody']),
            "X" if self.checkbox12_vars[i].get() else "",
            "X" if self.checkbox36_vars[i].get() else "",
            "N/A",]
            table_data.append(task_data)
            table = Table(table_data, colWidths=[300, 50, 50, 70])
            table.setStyle(TableStyle([ 
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'DejaVu'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),  # Smaller font size for table header
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'DejaVu'),
                ('FONTSIZE', (0, 1), (-1, -1), 6),  # Smaller font size for table rows
                ('GRID', (0, 0), (-1, -1), 1, colors.black),]))
            elements.append(table)
            elements.append(Spacer(7, 48))
            elements.append(Paragraph("Tarih: ____________________", styles['TurkishBody']))
            elements.append(Spacer(7, 12))
            elements.append(Paragraph("İsim Soyisim ve İmza", styles['TurkishBody']))
            pdf.build(elements)
    def create_form(self):
        self.create_form = tk.Tk()
        self.create_form.title("Bakım Kartı - 1000 Ton Pres")
        self.create_form.geometry("800x1000")
        main_frame = tk.Frame(self.create_form, padx=10, pady=10)
        main_frame.pack(fill="both", expand=True)
        # Create headers
        for idx, col in enumerate(["Bakım İşlemi", "12", "36", "Süre (dk)"]):
            label = tk.Label(main_frame, text=col, font=("Arial", 8, "bold"), borderwidth=1, relief="solid", padx=5, pady=5)
            label.grid(row=2, column=idx, sticky="nsew")
        # Create task rows
        for i, task in enumerate(self.tasks):
            task_label = tk.Label(main_frame, text=task, font=("Arial", 8), wraplength=600, justify="left", anchor="w", padx=5, pady=5)
            task_label.grid(row=i + 3, column=0, sticky="w", padx=5)            
            checkbox12_var = tk.BooleanVar()
            checkbox12 = ttk.Checkbutton(main_frame, variable=checkbox12_var)
            checkbox12.grid(row=i + 3, column=1, padx=5)
            self.checkbox12_vars.append(checkbox12_var)
            checkbox36_var = tk.BooleanVar()
            checkbox36 = ttk.Checkbutton(main_frame, variable=checkbox36_var)
            checkbox36.grid(row=i + 3, column=2, padx=5)
            self.checkbox36_vars.append(checkbox36_var)
            duration_entry = tk.Entry(main_frame, width=5)
            duration_entry.insert(0, str(5 + i * 5))
            duration_entry.grid(row=i + 3, column=3, padx=5)
            self.duration_entries.append(duration_entry)
        # Create save button
        save_button = ttk.Button(main_frame, text="PDF Kaydet", command=self.save_as_pdf)
        save_button.grid(row=len(self.tasks) + 4, columnspan=4, pady=10)
    def yüz_sayfası(self):
        self.view_kaynak_window = tk.Toplevel(self.main_window)
        self.view_kaynak_window.title("Yüz Sayfası")
        screen_width = self.main_window.winfo_screenwidth()
        screen_height = self.main_window.winfo_screenheight()
        window_width = int(screen_width)
        window_height = int(screen_height)
        window_x = (screen_width - window_width) // 2
        window_y = (screen_height - window_height) // 2
        self.view_kaynak_window.geometry(
            f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.view_kaynak_window.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")
        self.tree11 = ttk.Treeview(self.view_kaynak_window, columns=("Id", "Bakım Yapan", "Makine Adı", "Makine Kodu", "Bölümü", "Bakım Yapıldıgı Tarih", "Bir Sonraki Bakım Tarihi", "Çalışma Durum", "Makine Duruş Saati", "Makine Tekrar Çalısma Saati", "Resim", "Resim1"), show="headings", height=13)
        self.tree11.column("#0", width=0, stretch=tk.NO)
        self.tree11.heading("Id", text="Id")
        self.tree11.heading("Bakım Yapan", text="Bakım Yapan")
        self.tree11.heading("Makine Adı", text="Makine Adı")
        self.tree11.heading("Makine Kodu", text="Makine Kodu")
        self.tree11.heading("Bölümü", text="Bölümü")
        self.tree11.heading("Bakım Yapıldıgı Tarih", text="Bakım Yapıldıgı Tarih")
        self.tree11.heading("Bir Sonraki Bakım Tarihi", text="Bir Sonraki Bakım Tarihi")
        self.tree11.heading("Çalışma Durum", text="Çalışma Durum")
        self.tree11.heading("Makine Duruş Saati", text="Makine Duruş Saati")
        self.tree11.heading("Makine Tekrar Çalısma Saati", text="Makine Tekrar Çalısma Saati")
        self.tree11.heading("Resim", text="Resim")
        self.tree11.heading("Resim1", text="Resim1")
        # Configure columns
        self.tree11.column("Id", width=100)
        self.tree11.column("Bakım Yapan", width=100, anchor="center")
        self.tree11.column("Bakım Yapıldıgı Tarih", width=100, anchor="center")
        self.tree11.column("Makine Adı", width=120, anchor="center")
        self.tree11.column("Makine Kodu", width=100, anchor="center")
        self.tree11.column("Bölümü", width=90, anchor="center")
        self.tree11.column("Bir Sonraki Bakım Tarihi", width=90, anchor="center")
        self.tree11.column("Çalışma Durum", width=90, anchor="center")
        self.tree11.column("Makine Duruş Saati", width=90, anchor="center")
        self.tree11.column("Makine Tekrar Çalısma Saati", width=90, anchor="center")
        self.tree11.column("Resim", width=100, anchor="center")
        self.tree11.column("Resim1", width=100, anchor="center")
        self.tree11.pack(pady=80,padx=120)
        self.list_records11()
        self.tree11.pack(pady=80,padx=140)
        self.tree12 = ttk.Treeview(self.view_kaynak_window, columns=("Id", "Birim", "Adı", "Makine Adı", "Makine Kodu",
                                  "Bildirme Saati", "Bildirilen Arıza", "Arızalandıgı Yer", "Arıza Önceligi"), show="headings", height=13)
        self.tree12.heading("Id", text="Id")
        self.tree12.heading("Birim", text="Birim")
        self.tree12.heading("Adı", text="Adı")
        self.tree12.heading("Makine Adı", text="Makine Adı")
        self.tree12.heading("Makine Kodu", text="Makine Kodu")
        self.tree12.heading("Bildirme Saati", text="Bildirme Saati")
        self.tree12.heading("Bildirilen Arıza", text="Bildirilen Arıza")
        self.tree12.heading("Arızalandıgı Yer", text="Arzalandıgı yer")
        self.tree12.heading("Arıza Önceligi", text="Arıza Önceligi")
        self.tree12.column("#0", width=0, stretch=tk.NO)
        self.tree12.column("Id", width=100,anchor="center")
        self.tree12.column("Birim", width=100,anchor="center")
        self.tree12.column("Adı", width=100,anchor="center")
        self.tree12.column("Makine Adı", width=120,anchor="center")
        self.tree12.column("Makine Kodu", width=150,anchor="center")
        self.tree12.column("Bildirme Saati", width=150,anchor="center")
        self.tree12.column("Bildirilen Arıza", width=150,anchor="center")
        self.tree12.column("Arızalandıgı Yer",width=145,anchor="center")
        self.tree12.column("Arıza Önceligi", width=150,anchor="center")
        self.tree12.pack(pady=80,padx=140)
        resim6 = PhotoImage(file="C:/Users/admin/Downloads/bakım kartı.png")
        self.bakım_kart = tk.Button(self.view_kaynak_window,width=350,height=65,image=resim6,compound="left",borderwidth=1,relief="flat",command=self.create_form# Assuming you have a method for this
        )
        self.bakım_kart.resim6 = resim6
        self.bakım_kart.place(x=(screen_width)/90, y=175)
        # tk.Button(self.update_window, width=350, height=65, image=resim5,compound="left", borderwidth=1, relief="flat", command=self.update_güncelle)
        style = ttk.Style()
        # Adjust the row height as needed
        style.configure("Treeview", height=50)
     #########################-----------------------------------------------------------------------------------###############################
        self.view_kaynak_window.configure(background=self.bg_color)
        self.bakım_kart.configure(background=self.bg_color,activebackground=self.bg_color)
    ############################--------------------------------------------------------------------------------################################
    def list_records(self):
        conn = sqlite3.connect('user.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM makine")
        records = cursor.fetchall()  
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.tree.tag_configure("working", background="#90EE90")
        self.tree.tag_configure("not_working", background="#FF6347")
        for record in records:
            status = record[7]
            tag = "unknown"  # Default tag value
            if status in ["Calısıyor","Çalışıyor", "Calısıyor"]:
                tag = "working"
            elif status in ["Çalısmıyor", "Calısmıyor"]:
                tag = "not_working"
            self.tree.insert("", "end", values=record, tags=(tag,))
        conn.close()
    def list_record2(self):
        conn = sqlite3.connect("user.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM onarım_kayit")
        records = cursor.fetchall()
        for item in self.tree2.get_children():
            self.tree2.delete(item)
        for record in records:
            self.tree2.insert("", "end", values=(
                record[0], record[1], record[2], record[3], record[4], record[5], record[6]))
    def list_records7(self): 
       try: 
           conn = sqlite3.connect("user.db")
           cursor = conn.cursor()
           cursor.execute("SELECT Id, Parca_Adı FROM stok")
           stock_records = cursor.fetchall()
           stock_dict = {stock[0]: stock[1] for stock in stock_records}
           cursor.execute("SELECT * FROM stok_log")      
           records = cursor.fetchall()
           for item in self.tree6.get_children(): 
               self.tree6.delete(item)
           for record in records: 
               stock_id = record[0]
               stock_name = stock_dict.get(stock_id, "Unknown")
               miktar = record[2]
               if miktar <= 100: 
                   tag = 'red'
               elif 101 <= miktar <= 300: 
                   tag = 'orange'
               elif 301 <= miktar <= 500: 
                   tag = 'yellow'
               elif 501 <= miktar <= 1000: 
                   tag = 'lightgreen'
               else:
                   tag = 'green'
               self.tree6.insert("", "end", values=(stock_id, stock_name, record[1], record[2], record[3]), tags=(tag,))
       except sqlite3.Error as e: 
           print(f"An error occurred: {e}")
       finally: 
           if cursor: 
               cursor.close()
           if conn: 
               conn.close()
    def list_records9(self): 
         conn = sqlite3.connect('user.db')
         cursor = conn.cursor()
         cursor.execute("SELECT * FROM makine WHERE Bolum='PRES'")
         records = cursor.fetchall()  
         for item in self.tree8.get_children():
             self.tree8.delete(item)
         self.tree8.tag_configure("working", background="#90EE90")
         self.tree8.tag_configure("not_working", background="#FF6347")
         for record in records:
             status = record[7]
             tag = "unknown"  # Default tag value
             if status in ["Calısıyor","Çalışıyor", "Calısıyor"]:
                 tag = "working"
             elif status in ["Çalısmıyor", "Calısmıyor"]:
                 tag = "not_working"
             self.tree8.insert("", "end", values=record, tags=(tag,))
         conn.close()
    def list_records10(self):
        try:
            conn = sqlite3.connect("user.db")
        except sqlite3.Error:
            conn = sqlite3.connect("user.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM ariza_kayit WHERE Birim='PRES'")
        records = cursor.fetchall()
        for item in self.tree9.get_children():
            self.tree9.delete(item)
        for record in records:
            self.tree9.insert("", "end", values=(
                record[0], record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8]))
        self.tree9.tag_configure("high_priority", background="#FF6347")
        self.tree9.tag_configure("normal_priority", background="#FFD700")
        self.tree9.tag_configure("low_priority", background="#90EE90")
        for item in self.tree9.get_children():
            self.tree9.delete(item)
        for record in records:
            priority = record[8]
            if priority == "Yüksek":
                tag = "high_priority"
            elif priority == "Normal":
                tag = "normal_priority"
            else:
                tag = "low_priority"
            self.tree9.insert("", "end", values=record, tags=(tag,))
    def list_records11(self):
          conn = sqlite3.connect('user.db')
          cursor = conn.cursor()
          cursor.execute("SELECT * FROM makine WHERE Bolum='KAYNAK'")
          records = cursor.fetchall()  
          for item in self.tree11.get_children():
              self.tree11.delete(item)
          self.tree11.tag_configure("working", background="#90EE90")
          self.tree11.tag_configure("not_working", background="#FF6347")
          for record in records:
              status = record[7]
              tag = "unknown"  # Default tag value
              if status in ["Calısıyor","Çalışıyor", "Calısıyor"]:
                  tag = "working"
              elif status in ["Çalısmıyor", "Calısmıyor"]:
                  tag = "not_working"
              self.tree11.insert("", "end", values=record, tags=(tag,))
          conn.close()
    def list_records12(self):
        conn=sqlite3.connect('user.db')
        cursor=conn.cursor()
        cursor.execute("SELECT * FROM makine WHERE Bolum='TALAŞLI'")
        records=cursor.fetchall()
        for item in self.tree10.get_children():
            self.tree10.delete(item)
        self.tree10.tag_configure("working",background="#90EE90")
        self.tree10.tag_configure("not_working",background="#FF6347")
        for record in records:
            status=record[7]
            tag="unknown"
            if status in["Calısıyor","Çalışıyor","Calısıyor"]:
                tag="working"
            elif status in["Çalısmıyor","Calısmıyor"]:
                tag="not_working"
            self.tree10.insert("","end",values=record,tags=(tag,))
        conn.close()
    def list_records13(self):
       
            conn=sqlite3.connect("user.db")
            cursor=conn.cursor()
            cursor.execute("SELECT * FROM ariza_kayit WHERE Birim='PRES'")
            records=cursor.fetchall()
            for item in self.tree16.get_children():
                self.tree16.delete(item)
            for record in records:
                self.tree16.insert("","end",values=(record[0],record[1],record[2],record[3],record[4],record[5],record[6],record[7],record[8]))
                self.tree16.tag_configure("high_priority",background="#FF6347")
                self.tree16.tag_configure("normal_priority",background="#FFD700")
                self.tree16.tag_configure("low_priority",background="#90EE90")
                for item in self.tree16.get_children():
                    self.tree16.delete(item)
                for record in records:
                    priority=record[8]
                    if priority =="Yüksek":
                        tag="high_priority"
                    elif priority =="Normal":
                        tag="normal_priority"
                    else:
                        tag="low_priority"
                    self.tree16.insert("","end",values=record,tags=(tag,))
    def list_records14(self):
         conn=sqlite3.connect('user.db')
         cursor=conn.cursor()
         cursor.execute("SELECT * FROM makine")
         records=cursor.fetchall()
         for item in self.tree17.get_children():
             self.tree17.delete(item)
         self.tree17.tag_configure("working",background="#90EE90")
         self.tree17.tag_configure("not_working",background="#FF6347")
         for record in records:
             status=record[7]
             tag="unknown"
             if status in["Calısıyor","Çalışıyor","Calısıyor"]:
                 tag="working"
             elif status in["Çalısmıyor","Calısmıyor"]:
                 tag="not_working"
             self.tree17.insert("","end",values=record,tags=(tag,))
         conn.close()
    def list_records15(self):
        conn=sqlite3.connect("user.db")
        cursor=conn.cursor()
        cursor.execute("SELECT * FROM ariza_kayit")
        records=cursor.fetchall()
        for item in self.tree18.get_children():
            self.tree18.delete(item)
        for record in records:
            self.tree18.insert("","end",values=(record[0],record[1],record[2],record[3],record[4],record[5],record[6],record[7],record[8]))
            self.tree18.tag_configure("high_priority",background="#FF6347")
            self.tree18.tag_configure("normal_priority",background="#FFD700")
            self.tree18.tag_configure("low_priority",background="#90EE90")
            for item in self.tree18.get_children():
                self.tree18.delete(item)
            for record in records:
                priority=record[8]
                if priority =="Yüksek":
                    tag="high_priority"
                elif priority =="Normal":
                    tag="normal_priority"
                else:
                    tag="low_priority"
                self.tree18.insert("","end",values=record,tags=(tag,))
        
    def list_records3(self):
        try:
            conn = sqlite3.connect("user.db")
        except sqlite3.Error:
            conn = sqlite3.connect("user.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM ariza_kayit")
        records = cursor.fetchall()
        for item in self.tree1.get_children():
            self.tree1.delete(item)
        for record in records:
            self.tree1.insert("", "end", values=(
                record[0], record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8]))
        self.tree1.tag_configure("high_priority", background="#FF6347")
        self.tree1.tag_configure("normal_priority", background="#FFD700")
        self.tree1.tag_configure("low_priority", background="#90EE90")
        for item in self.tree1.get_children():
            self.tree1.delete(item)
        for record in records:
            priority = record[8]
            if priority == "Yüksek":
                tag = "high_priority"
            elif priority == "Normal":
                tag = "normal_priority"
            else:
                tag = "low_priority"
            self.tree1.insert("", "end", values=record, tags=(tag,))
    def list_records4(self):
        conn = sqlite3.connect("user.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tasks")
        records = cursor.fetchall()
        conn.close()
        for item in self.tree3.get_children():
           self.tree3.delete(item)
        for record in records:
           status = record[6]
           if status in ["Çalısıyor", "Çalışıyor"]:
               tag = 'working'
           elif status in ["Çalısmıyor", "Çalışmıyor"]:
               tag = 'not_working'
           elif status == "Molada":
               tag = 'on_break'
           else:
               tag = ''
           self.tree3.insert("", "end", values=record, tags=(tag,))
    def list_records7(self):
        conn=sqlite3.connect("user.db")
        cursor=conn.cursor()
        cursor.execute("SELECT * FROM users")
        records=cursor.fetchall()
        """
        for item in self.tree15.get_children():
            self.tree15.delete(item)
        for record in records:
            self.tree15.insert("","end",values=(record[0],record[1],record[2],record[3]))
            """
    def list_records6(self):
        conn = sqlite3.connect("user.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM stok")
        records = cursor.fetchall()
        conn.close()
        for item in self.tree5.get_children():
            self.tree5.delete(item)  
        for record in records:
            miktar = record[4]
            if miktar <= 100:
                tag = 'red'
            elif 100 <= miktar <= 300:
                tag = 'orange'
            elif 301 <= miktar <= 500:
                tag = 'yellow'
            elif 501 <= miktar <= 1000:
                tag = 'lightgreen'
            else:
                tag = 'green'
            self.tree5.insert("", "end", values=(record[0], record[1], record[2], record[3], miktar), tags=(tag,))
############################--------------------------------------------------------------------------------################################
    def create_popup_menu5(self, event): 
        self.popup_menu5 = tk.Menu(self.main_window, tearoff=0)
        self.popup_menu5.add_command(label="Molada", command=self.set_worker_on_break)
        self.popup_menu5.add_command(label="Çalısıyor",command=self.set_worker_on_continue)
        self.popup_menu5.add_command(label="İş Bitis",command=self.set_worker_on_break1)
    # Display the popup menu
        try: 
            self.popup_menu5.tk_popup(event.x_root, event.y_root)
        finally: 
            self.popup_menu5.grab_release()
    def create_popup_menu3(self, event):
        popup_menu = tk.Menu(tearoff=0)
        popup_menu.add_command(label="Talep oluştur", command=self.talep_oluştur)
        popup_menu.add_command(label="Makine Durdur", command=self.makine_durdur)
        popup_menu.add_command(label="Makine Başlat", command=self.makine_baslat)
        popup_menu.add_command(label="Makine Güncelle", command=self.makine_guncelle)
        image_menu = tk.Menu(popup_menu, tearoff=0)
        image_menu.add_command(label="1. Resim", command=lambda: self.find_image_page(11))
        image_menu.add_command(label="2. Resim", command=lambda: self.find_image_page(12))
        image_menu.add_command(label="3.Resim ", command=lambda:self.find_image_page(13))
        image_menu.add_command(label="4.Resim",command=lambda:self.find_image_page(14))
        image_menu.add_command(label="5.Resim",command=lambda:self.find_image_page(15))
        popup_menu.add_cascade(label="Resmi Görüntüle", menu=image_menu)
        popup_menu.post(event.x_root, event.y_root)
    def create_popup_menu6(self,event):
        popup_menu=tk.Menu(tearoff=0)
        popup_menu.add_command(label="Pdf Aktar",command=self.export_to_pdf1)
        popup_menu.post(event.x_root, event.y_root)
    def create_popup_menu7(self,event):
        popup_menu=tk.Menu(tearoff=0)
        popup_menu.add_command(label="Depoda Ürün Çıkart",command=self.stok_azalt)
        popup_menu.add_command(label="Depoya ürün Ekle",command=self.stok_ekle)
        popup_menu.post(event.x_root,event.y_root)
    def create_popup_menu4(self, event): 
        popup_menu = tk.Menu(tearoff=0)
        popup_menu.add_command(label="Talep Bakımı", command=self.onarım_page)
        popup_menu.add_command(label="Pdf Aktar",command=self.export_to_pdf)
        popup_menu.add_command(label="Görev Ata",command=self.görev_ata)
        popup_menu.post(event.x_root, event.y_root)
    # Call a method to populate the Treeviws with data (e.g., self.list_records())
app = FabricaApp()
app.login_window.mainloop()
